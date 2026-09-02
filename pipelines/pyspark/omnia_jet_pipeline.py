import os
import sys
import json
import re
import datetime
import math
import argparse
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from common_utils import clean_str, parse_num, parse_date_str, parse_date_obj, date_to_iso, log_event

# Polars high-speed multithreaded local offline acceleration engine
try:
    import polars as pl
    POLARS_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False

def run_omnia_jet(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    run_id = config.get('runId', 'UNKNOWN_RUN')
    run_dir = os.path.dirname(os.path.dirname(config_path))
    output_dir = os.path.join(run_dir, 'output')
    os.makedirs(output_dir, exist_ok=True)
    log_file = os.path.join(run_dir, 'logs', 'execution.txt')
    
    log_event(run_id, 'INITIALIZATION', 5, 'Initializing Omnia JET Workflow & Fraud Analytics engine', log_file)
    if POLARS_AVAILABLE:
        log_event(run_id, 'ACCELERATION', 8, 'Polars Multi-Threaded Acceleration Activated (100% Offline Local Compute)', log_file)

    input_files = config.get('files', [])
    dataset_map = config.get('datasetMap', {})
    field_mappings = config.get('fieldMappings', {})
    params = config.get('omniaParameters', {}) or {}

    log_event(run_id, 'READ_INPUT', 10, f'Reading Omnia input datasets ({len(input_files)} uploaded)', log_file)

    def load_dataset(file_id, sheet_name=None, dataset_type=None):
        cache_dir = os.path.join(run_dir, 'cache')
        if dataset_type and os.path.exists(cache_dir):
            parquet_path = os.path.join(cache_dir, f"{dataset_type}.parquet")
            if os.path.exists(parquet_path):
                if POLARS_AVAILABLE:
                    try:
                        return pl.read_parquet(parquet_path).to_pandas().astype(str)
                    except Exception:
                        pass
                try:
                    return pd.read_parquet(parquet_path).astype(str)
                except Exception:
                    pass

        for f in input_files:
            if f.get('fileId') == file_id or f.get('fileName') == file_id:
                path = f.get('filePath')
                ext = f.get('extension', '').lower()
                if ext in ('xlsx', 'xls'):
                    if sheet_name:
                        return pd.read_excel(path, sheet_name=sheet_name, dtype=str)
                    return pd.read_excel(path, dtype=str)
                else:
                    if POLARS_AVAILABLE:
                        try:
                            return pl.read_csv(path, infer_schema_length=0, truncate_ragged_lines=True).to_pandas()
                        except Exception:
                            pass
                    try:
                        return pd.read_csv(path, dtype=str, skipinitialspace=True)
                    except:
                        return pd.read_csv(path, sep=None, engine='python', dtype=str)
        return None

    tb_df = None
    tb_beg_df = None   # Prior-year (beginning) TB
    tb_end_df = None   # Current-year (ending) TB
    gl_df = None
    coa_df = None

    # Sheet-name aliases for beginning / ending TB
    BEG_SHEET_NAMES = {'tb_beginning', 'tbbeg', 'tb_beg', 'beginning', 'tb_prior'}
    END_SHEET_NAMES = {'tb_ending', 'tbend', 'tb_end', 'ending', 'tb_current'}

    for f in input_files:
        detected = f.get('detectedDataset')
        if detected == 'TRIAL_BALANCE' and tb_df is None:
            tb_df = load_dataset(f.get('fileId'), dataset_type='tb')
        elif detected in ('GENERAL_LEDGER', 'POPULATION') and gl_df is None:
            gl_df = load_dataset(f.get('fileId'), dataset_type='gl')
        elif detected == 'COA' and coa_df is None:
            coa_df = load_dataset(f.get('fileId'), dataset_type='coa')
        elif f.get('sheets'):
            for s in f.get('sheets', []):
                s_class = s.get('detectedDataset')
                s_name  = (s.get('sheetName') or '').lower().replace(' ', '_')
                if s_class in ('TRIAL_BALANCE', 'TRIAL_BALANCE_BEG') or s_name in BEG_SHEET_NAMES:
                    if tb_beg_df is None:
                        tb_beg_df = load_dataset(f.get('fileId'), s.get('sheetName'))
                    elif tb_df is None:
                        tb_df = load_dataset(f.get('fileId'), s.get('sheetName'), dataset_type='tb')
                elif s_class == 'TRIAL_BALANCE_END' or s_name in END_SHEET_NAMES:
                    if tb_end_df is None:
                        tb_end_df = load_dataset(f.get('fileId'), s.get('sheetName'))
                elif s_class in ('GENERAL_LEDGER', 'POPULATION') and gl_df is None:
                    gl_df = load_dataset(f.get('fileId'), s.get('sheetName'), dataset_type='gl')
                elif s_class == 'COA' and coa_df is None:
                    coa_df = load_dataset(f.get('fileId'), s.get('sheetName'), dataset_type='coa')

    if dataset_map.get('tbFileId'):
        tb_df = load_dataset(dataset_map['tbFileId'], dataset_map.get('tbSheetName'), dataset_type='tb')
    if dataset_map.get('glFileId'):
        gl_df = load_dataset(dataset_map['glFileId'], dataset_map.get('glSheetName'), dataset_type='gl')
    if dataset_map.get('coaFileId'):
        coa_df = load_dataset(dataset_map['coaFileId'], dataset_map.get('coaSheetName'), dataset_type='coa')

    # Merge Beginning + Ending TB if both are present
    if tb_beg_df is not None and tb_end_df is not None:
        log_event(run_id, 'TB_MERGE', 12,
                  f'Merging TB_Beginning ({len(tb_beg_df)} rows) + TB_Ending ({len(tb_end_df)} rows) into unified TB',
                  log_file)
        all_cols = list(dict.fromkeys(list(tb_beg_df.columns) + list(tb_end_df.columns)))
        tb_beg_df = tb_beg_df.reindex(columns=all_cols)
        tb_end_df = tb_end_df.reindex(columns=all_cols)
        tb_df = pd.concat([tb_beg_df, tb_end_df], ignore_index=True)
    elif tb_beg_df is not None and tb_df is None:
        tb_df = tb_beg_df
    elif tb_end_df is not None and tb_df is None:
        tb_df = tb_end_df

    if tb_df is None:
        raise ValueError("Trial Balance (TB) dataset is required for Omnia JET workflow.")
    if gl_df is None:
        raise ValueError("General Ledger Detail (JE) dataset is required for Omnia JET workflow.")
    if coa_df is None:
        raise ValueError("Chart of Accounts (COA) dataset is required for Omnia JET workflow.")

    log_event(run_id, 'MAPPING', 18, 'Applying Omnia CDM standardized field mappings', log_file)

    def apply_mapping(df, mapping_list):
        if not mapping_list:
            return df
        rename_map = {}
        for m in mapping_list:
            std = m.get('standardField')
            src = m.get('sourceField')
            if src and src in df.columns:
                rename_map[src] = std
        return df.rename(columns=rename_map)

    tb_df = apply_mapping(tb_df, field_mappings.get('tb', []))
    gl_df = apply_mapping(gl_df, field_mappings.get('gl', []))
    coa_df = apply_mapping(coa_df, field_mappings.get('coa', []))

    def get_col(df, target_names, default=""):
        for name in target_names:
            if name in df.columns:
                return df[name]
        return pd.Series([default] * len(df), index=df.index)

    dec_sep = params.get('decimalSeparator', 'Period')

    # -------------------------------------------------------------
    # 1. PREPARING GENERAL LEDGER DETAIL
    # -------------------------------------------------------------
    log_event(run_id, 'GL_PREPARATION', 25, 'Transforming and standardizing General Ledger Detail', log_file)

    gl_clean = pd.DataFrame()
    gl_clean['entity_id'] = get_col(gl_df, ['entity_id', 'entity', 'Entity Number', 'Company Code']).apply(clean_str)
    gl_clean['entity_name'] = get_col(gl_df, ['entity_name', 'company name']).apply(clean_str)
    gl_clean['journal_number'] = get_col(gl_df, ['journal_number', 'documentno', 'Document number', 'Accounting document']).apply(clean_str)
    gl_clean['journal_line_number'] = get_col(gl_df, ['journal_line_number', 'line', 'item', 'Line number']).apply(clean_str)
    gl_clean['account_number'] = get_col(gl_df, ['account_number', 'g_l', 'gl', 'GL Account', 'Account Number']).apply(clean_str)
    gl_clean['account_description'] = get_col(gl_df, ['account_description', 'description', 'GL Description']).apply(clean_str)
    gl_clean['journal_header_description'] = get_col(gl_df, ['journal_header_description', 'document_header_text', 'Text Header']).apply(clean_str)
    gl_clean['journal_line_description'] = get_col(gl_df, ['journal_line_description', 'text', 'Text Details']).apply(clean_str)
    
    gl_clean['date_effective'] = get_col(gl_df, ['date_effective', 'entry_date', 'Accounting date']).apply(parse_date_str)
    gl_clean['date_posted'] = get_col(gl_df, ['date_posted', 'pstng_date', 'Posting date']).apply(parse_date_str)
    gl_clean['fiscal_year'] = get_col(gl_df, ['fiscal_year', 'year']).apply(clean_str)
    gl_clean['fiscal_period'] = get_col(gl_df, ['fiscal_period', 'period']).apply(clean_str)
    
    gl_clean['net_amount_ec'] = get_col(gl_df, ['net_amount_ec', 'amount_in_local_cur', 'Amount in local currency', 'Amount']).apply(lambda x: parse_num(x, dec_sep))
    gl_clean['entity_currency_ec'] = get_col(gl_df, ['entity_currency_ec', 'lcurr', 'Local Currency', 'Curr', 'Currency code'], 'INR').apply(clean_str)
    
    raw_debit_ec = get_col(gl_df, ['debit_amount_ec', 'debit']).apply(lambda x: parse_num(x, dec_sep))
    raw_credit_ec = get_col(gl_df, ['credit_amount_ec', 'credit']).apply(lambda x: parse_num(x, dec_sep))
    
    gl_clean['debit_amount_ec'] = np.where((raw_debit_ec == 0) & (raw_credit_ec == 0) & (gl_clean['net_amount_ec'] > 0), gl_clean['net_amount_ec'], raw_debit_ec)
    gl_clean['credit_amount_ec'] = np.where((raw_debit_ec == 0) & (raw_credit_ec == 0) & (gl_clean['net_amount_ec'] < 0), -gl_clean['net_amount_ec'], raw_credit_ec)

    gl_clean['net_amount_gc'] = get_col(gl_df, ['net_amount_gc', 'amount_in_loc_curr_2', 'Amount in group curr']).apply(lambda x: parse_num(x, dec_sep))
    gl_clean['group_currency_gc'] = get_col(gl_df, ['group_currency_gc', 'lcur2', 'Group Currency']).apply(clean_str)
    gl_clean['debit_amount_gc'] = get_col(gl_df, ['debit_amount_gc']).apply(lambda x: parse_num(x, dec_sep))
    gl_clean['credit_amount_gc'] = get_col(gl_df, ['credit_amount_gc']).apply(lambda x: parse_num(x, dec_sep))

    gl_clean['net_amount_oc'] = get_col(gl_df, ['net_amount_oc', 'amount_in_doc_curr', 'Amount in doc curr']).apply(lambda x: parse_num(x, dec_sep))
    gl_clean['original_currency_oc'] = get_col(gl_df, ['original_currency_oc', 'curr', 'Doc Curr']).apply(clean_str)
    gl_clean['debit_amount_oc'] = get_col(gl_df, ['debit_amount_oc']).apply(lambda x: parse_num(x, dec_sep))
    gl_clean['credit_amount_oc'] = get_col(gl_df, ['credit_amount_oc']).apply(lambda x: parse_num(x, dec_sep))

    gl_clean['transaction_type'] = get_col(gl_df, ['transaction_type', 'type', 'Document type']).apply(clean_str)
    gl_clean['transaction_type_description'] = get_col(gl_df, ['transaction_type_description']).apply(clean_str)
    gl_clean['time_posted'] = get_col(gl_df, ['time_posted', 'time']).apply(clean_str)
    gl_clean['userid_entered'] = get_col(gl_df, ['userid_entered', 'user_name', 'User UD', 'User Name', 'Username']).apply(clean_str)
    gl_clean['user_name_entered'] = get_col(gl_df, ['user_name_entered', 'user_name', 'User Name']).apply(clean_str)
    
    raw_is_std = get_col(gl_df, ['is_standard', 'standard_flag', 'mansys']).apply(clean_str).str.upper()
    gl_clean['is_standard'] = np.where(raw_is_std.isin(['S', 'STANDARD', 'Y', 'YES', '1']), 'S',
                              np.where(raw_is_std.isin(['N', 'NON-STANDARD', 'NONSTANDARD', 'NO', '0']), 'N',
                              np.where(gl_clean['transaction_type'].isin(['SA', 'AB', 'MANUAL']), 'N', 'S')))

    gl_clean['source'] = get_col(gl_df, ['source', 'source_system']).apply(clean_str)
    gl_clean['dc_indicator'] = np.where(gl_clean['net_amount_ec'] >= 0, 'D', 'C')

    if (gl_clean['entity_id'] == '').all():
        gl_clean['entity_id'] = 'ENT01'

    gl_out_csv = os.path.join(output_dir, 'General_Ledger_Detail.csv')
    gl_clean.to_csv(gl_out_csv, index=False)

    # -------------------------------------------------------------
    # 2. PREPARING TRIAL BALANCE
    # -------------------------------------------------------------
    log_event(run_id, 'TB_PREPARATION', 35, 'Transforming and standardizing Trial Balance', log_file)

    testing_period_start = params.get('testingPeriodStart') or params.get('first_day_testing_period') or '04/01/2025'
    testing_period_end = params.get('testingPeriodEnd') or params.get('last_day_testing_period') or params.get('fiscalYearEnd') or '03/31/2026'

    start_date_obj = parse_date_obj(testing_period_start) or datetime.date(2025, 4, 1)
    end_date_obj = parse_date_obj(testing_period_end) or datetime.date(2026, 3, 31)

    prior_period_end_date = (start_date_obj - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
    current_period_end_date = end_date_obj.strftime('%Y-%m-%d')
    current_fy = str(params.get('fiscalYear') or 2026)
    try:
        prior_fy = str(int(current_fy) - 1)
    except:
        prior_fy = "2025"

    tb_base = pd.DataFrame()
    tb_base['entity_id'] = get_col(tb_df, ['entity_id', 'entity', 'Entity Number', 'Company Code']).apply(clean_str)
    tb_base['entity_name'] = get_col(tb_df, ['entity_name', 'company name']).apply(clean_str)
    tb_base['account_number'] = get_col(tb_df, ['account_number', 'g_l', 'gl', 'GL Account', 'Account Number']).apply(clean_str)
    tb_base['account_description'] = get_col(tb_df, ['account_description', 'description', 'GL Description']).apply(clean_str)
    tb_base['raw_period_end_date'] = get_col(tb_df, ['period_end_date', 'end_date', 'date', 'Closing Date']).apply(date_to_iso)
    tb_base['entity_currency_ec'] = get_col(tb_df, ['entity_currency_ec', 'currency', 'lcurr', 'Local Currency'], 'INR').apply(clean_str)
    tb_base['group_currency_gc'] = get_col(tb_df, ['group_currency_gc', 'group_currency'], 'USD').apply(clean_str)
    
    tb_base['beginning_balance_ec'] = get_col(tb_df, ['beginning_balance_ec', 'opening_balance', 'Opening Balance', 'beginning_balance', 'op balance']).apply(lambda x: parse_num(x, dec_sep))
    tb_base['ending_balance_ec'] = get_col(tb_df, ['ending_balance_ec', 'closing_balance', 'Closing Balance', 'ending_balance', 'balance']).apply(lambda x: parse_num(x, dec_sep))
    tb_base['beginning_balance_gc'] = get_col(tb_df, ['beginning_balance_gc']).apply(lambda x: parse_num(x, dec_sep))
    tb_base['ending_balance_gc'] = get_col(tb_df, ['ending_balance_gc']).apply(lambda x: parse_num(x, dec_sep))
    
    tb_base['beginning_balance_gc'] = np.where(tb_base['beginning_balance_gc'] == 0, (tb_base['beginning_balance_ec'] / 83.1).round(2), tb_base['beginning_balance_gc'])
    tb_base['ending_balance_gc'] = np.where(tb_base['ending_balance_gc'] == 0, (tb_base['ending_balance_ec'] / 83.1).round(2), tb_base['ending_balance_gc'])

    tb_base['period_type'] = get_col(tb_df, ['period_type'], 'YTD').apply(clean_str)
    tb_base['chart_of_accounts'] = get_col(tb_df, ['chart_of_accounts'], 'DEFAULT').apply(clean_str)

    if (tb_base['entity_id'] == '').all():
        tb_base['entity_id'] = gl_clean['entity_id'].iloc[0] if len(gl_clean) > 0 else 'ENT01'

    distinct_dates = set(d for d in tb_base['raw_period_end_date'].unique() if d)
    has_dual_dates = (prior_period_end_date in distinct_dates and current_period_end_date in distinct_dates) or len(distinct_dates) >= 2

    if has_dual_dates:
        tb_clean = tb_base.copy()
        tb_clean['period_end_date'] = tb_clean['raw_period_end_date']
        tb_clean['Fiscal_Year_Identifier'] = np.where(
            tb_clean['period_end_date'] <= prior_period_end_date, 'Current Period Beginning', 'Current Period Ending'
        )
        tb_clean['fiscal_year'] = np.where(
            tb_clean['Fiscal_Year_Identifier'] == 'Current Period Beginning', prior_fy, current_fy
        )
        tb_clean['fiscal_period'] = '12'
    else:
        tb_beg = tb_base.copy()
        tb_beg['period_end_date'] = prior_period_end_date
        tb_beg['ending_balance_ec'] = tb_base['beginning_balance_ec']
        tb_beg['ending_balance_gc'] = tb_base['beginning_balance_gc']
        tb_beg['beginning_balance_ec'] = 0.0
        tb_beg['beginning_balance_gc'] = 0.0
        tb_beg['fiscal_year'] = prior_fy
        tb_beg['fiscal_period'] = '12'
        tb_beg['Fiscal_Year_Identifier'] = 'Current Period Beginning'

        tb_end = tb_base.copy()
        tb_end['period_end_date'] = current_period_end_date
        tb_end['ending_balance_ec'] = tb_base['ending_balance_ec']
        tb_end['ending_balance_gc'] = tb_base['ending_balance_gc']
        tb_end['beginning_balance_ec'] = tb_base['beginning_balance_ec']
        tb_end['beginning_balance_gc'] = tb_base['beginning_balance_gc']
        tb_end['fiscal_year'] = current_fy
        tb_end['fiscal_period'] = '12'
        tb_end['Fiscal_Year_Identifier'] = 'Current Period Ending'

        tb_beg.drop(columns=['raw_period_end_date'], errors='ignore').to_csv(os.path.join(output_dir, 'TB_Start.csv'), index=False)
        tb_end.drop(columns=['raw_period_end_date'], errors='ignore').to_csv(os.path.join(output_dir, 'TB_End.csv'), index=False)

        tb_clean = pd.concat([tb_beg, tb_end], ignore_index=True)

    tb_clean = tb_clean.drop(columns=['raw_period_end_date'], errors='ignore')
    tb_out_csv = os.path.join(output_dir, 'Trial_Balance.csv')
    tb_clean.to_csv(tb_out_csv, index=False)

    # -------------------------------------------------------------
    # 3. PREPARING CHART OF ACCOUNTS (COA)
    # -------------------------------------------------------------
    log_event(run_id, 'COA_PREPARATION', 42, 'Transforming and standardizing Chart of Accounts', log_file)

    coa_clean = pd.DataFrame()
    coa_clean['chart_of_accounts'] = get_col(coa_df, ['chart_of_accounts'], 'DEFAULT').apply(clean_str)
    coa_clean['account_number'] = get_col(coa_df, ['account_number', 'g_l', 'gl', 'GL Account', 'Account Number']).apply(clean_str)
    coa_clean['account_description'] = get_col(coa_df, ['account_description', 'description', 'GL Description']).apply(clean_str)
    
    raw_cat = get_col(coa_df, ['financial_statement_category', 'fs category', 'category', 'account_subtype']).apply(clean_str).str.lower()
    coa_clean['financial_statement_category'] = np.where(raw_cat.str.contains('asset'), 'Assets',
                                                np.where(raw_cat.str.contains('liab'), 'Liabilities',
                                                np.where(raw_cat.str.contains('equity'), 'Equity',
                                                np.where(raw_cat.str.contains('rev') | raw_cat.str.contains('income'), 'Revenue',
                                                np.where(raw_cat.str.contains('exp'), 'Expenses', 'Assets')))))
    
    coa_clean['financial_statement_line'] = get_col(coa_df, ['financial_statement_line', 'fs_line_item', 'fs line']).apply(clean_str)
    coa_clean['account_grouping_1'] = get_col(coa_df, ['account_grouping_1', 'grouping']).apply(clean_str)
    coa_clean['financial_statement_type'] = np.where(coa_clean['financial_statement_category'].isin(['Assets', 'Liabilities', 'Equity']), 'BS', 'IS')
    coa_clean['entity_id'] = get_col(coa_df, ['entity_id'], tb_clean['entity_id'].iloc[0] if len(tb_clean) > 0 else 'ENT01').apply(clean_str)
    coa_clean['revision_date'] = get_col(coa_df, ['revision_date']).apply(parse_date_str)

    coa_out_csv = os.path.join(output_dir, 'Chart_of_Accounts.csv')
    coa_clean.to_csv(coa_out_csv, index=False)

    # -------------------------------------------------------------
    # 4. OMNIA RECONCILIATION
    # -------------------------------------------------------------
    log_event(run_id, 'RECONCILIATION', 50, 'Performing account-level reconciliation (TB vs JE Activity)', log_file)

    je_summary = gl_clean.groupby(['entity_id', 'account_number']).agg(
        je_activity=('net_amount_ec', 'sum'),
        debit_amount=('debit_amount_ec', 'sum'),
        credit_amount=('credit_amount_ec', 'sum'),
        number_of_lines=('journal_line_number', 'count'),
        number_of_entries=('journal_number', 'nunique')
    ).reset_index()

    tb_beg_sub = tb_clean[tb_clean['Fiscal_Year_Identifier'] == 'Current Period Beginning'][
        ['entity_id', 'account_number', 'account_description', 'ending_balance_ec', 'ending_balance_gc']
    ].rename(columns={'ending_balance_ec': 'beginning_balance', 'ending_balance_gc': 'beginning_balance_gc'})

    tb_end_sub = tb_clean[tb_clean['Fiscal_Year_Identifier'] == 'Current Period Ending'][
        ['entity_id', 'account_number', 'account_description', 'ending_balance_ec', 'ending_balance_gc']
    ].rename(columns={'ending_balance_ec': 'ending_balance', 'ending_balance_gc': 'ending_balance_gc'})

    if len(tb_beg_sub) > 0 and len(tb_end_sub) > 0:
        tb_reconciled_base = tb_end_sub.merge(
            tb_beg_sub[['entity_id', 'account_number', 'beginning_balance', 'beginning_balance_gc']],
            on=['entity_id', 'account_number'],
            how='outer'
        )
    elif len(tb_end_sub) > 0:
        tb_reconciled_base = tb_end_sub.copy()
        tb_reconciled_base['beginning_balance'] = 0.0
        tb_reconciled_base['beginning_balance_gc'] = 0.0
    else:
        tb_reconciled_base = tb_clean[['entity_id', 'account_number', 'account_description', 'beginning_balance_ec', 'ending_balance_ec']].rename(
            columns={'beginning_balance_ec': 'beginning_balance', 'ending_balance_ec': 'ending_balance'}
        ).drop_duplicates(subset=['entity_id', 'account_number'])

    tb_reconciled_base['beginning_balance'] = tb_reconciled_base['beginning_balance'].fillna(0.0)
    tb_reconciled_base['ending_balance'] = tb_reconciled_base['ending_balance'].fillna(0.0)

    recon_df = tb_reconciled_base.merge(je_summary, on=['entity_id', 'account_number'], how='outer')
    recon_df = recon_df.merge(
        coa_clean[['account_number', 'financial_statement_category', 'financial_statement_line', 'account_grouping_1']],
        on='account_number',
        how='left'
    )

    recon_df['beginning_balance'] = recon_df['beginning_balance'].fillna(0.0)
    recon_df['ending_balance'] = recon_df['ending_balance'].fillna(0.0)
    recon_df['je_activity'] = recon_df['je_activity'].fillna(0.0)
    recon_df['debit_amount'] = recon_df['debit_amount'].fillna(0.0)
    recon_df['credit_amount'] = recon_df['credit_amount'].fillna(0.0)
    recon_df['number_of_lines'] = recon_df['number_of_lines'].fillna(0).astype(int)
    recon_df['number_of_entries'] = recon_df['number_of_entries'].fillna(0).astype(int)

    recon_df['trial_activity'] = recon_df['ending_balance'] - recon_df['beginning_balance']
    recon_df['variance'] = recon_df['ending_balance'] - recon_df['beginning_balance'] - recon_df['je_activity']
    recon_df['abs_variance'] = recon_df['variance'].abs()

    recon_df['reconciliation'] = np.where(recon_df['abs_variance'] <= 1.0, 'Reconciled', 'Unreconciled')
    recon_df['reconciliation_currency'] = 'Entity Currency'

    recon_out_csv = os.path.join(output_dir, 'Parquet_Reconciliation.csv')
    recon_df.to_csv(recon_out_csv, index=False)

    unrecon_df = recon_df[recon_df['reconciliation'] == 'Unreconciled'].copy()
    unrecon_out_csv = os.path.join(output_dir, 'Unreconciled_Accounts_Detail.csv')
    unrecon_df.to_csv(unrecon_out_csv, index=False)

    reconciled_count = int((recon_df['reconciliation'] == 'Reconciled').sum())
    unreconciled_count = int((recon_df['reconciliation'] == 'Unreconciled').sum())

    # -------------------------------------------------------------
    # 5. DATA INTEGRITY CHECKS (ALL 20 DQCs)
    # -------------------------------------------------------------
    log_event(run_id, 'DQC_EXECUTION', 60, 'Executing all 20 Data Quality Checks (DQC 01a to 20)', log_file)

    dqc_results = []
    coa_accounts = set(coa_clean['account_number'].unique())
    tb_accounts = set(tb_clean['account_number'].unique())

    # 01a: COA Blank Values
    coa_blank = coa_clean[(coa_clean['account_number'] == '') | (coa_clean['account_description'] == '') | (coa_clean['financial_statement_category'] == '')]
    dqc_results.append({
        'check': '01a_Error_COA_Blank_Values', 'desc': 'Chart of accounts has blank values in critical fields',
        'type': 'Error', 'capability': 'Core Mapping', 'fields': 'account_number, account_description, financial_statement_category',
        'affected_je': 0, 'affected_lines': len(coa_blank), 'debit_ec': 0.0, 'credit_ec': 0.0
    })

    # 01b: TB Blank Values
    tb_blank = tb_clean[(tb_clean['entity_id'] == '') | (tb_clean['account_number'] == '') | (tb_clean['period_end_date'] == '')]
    dqc_results.append({
        'check': '01b_Error_TB_Blank_Values', 'desc': 'Trial balance has blank values in critical fields',
        'type': 'Error', 'capability': 'Core Mapping', 'fields': 'entity_id, account_number, period_end_date',
        'affected_je': 0, 'affected_lines': len(tb_blank), 'debit_ec': 0.0, 'credit_ec': 0.0
    })

    # 01c: JE Blank Values
    je_blank = gl_clean[(gl_clean['entity_id'] == '') | (gl_clean['journal_number'] == '') | (gl_clean['account_number'] == '') | (gl_clean['date_effective'] == '')]
    dqc_results.append({
        'check': '01c_Error_JE_Blank_Values', 'desc': 'Journal entry data has blank values in critical fields',
        'type': 'Error', 'capability': 'Core Mapping', 'fields': 'entity_id, journal_number, account_number, date_effective',
        'affected_je': je_blank['journal_number'].nunique(), 'affected_lines': len(je_blank),
        'debit_ec': je_blank['debit_amount_ec'].sum(), 'credit_ec': je_blank['credit_amount_ec'].sum()
    })

    # 01d: JE Blank User ID
    je_blank_user = gl_clean[gl_clean['userid_entered'] == '']
    dqc_results.append({
        'check': '01d_Warning_JE_Blank_UserID', 'desc': 'Journal entry data has blank values in User ID',
        'type': 'Warning', 'capability': 'User Analytics', 'fields': 'userid_entered',
        'affected_je': je_blank_user['journal_number'].nunique(), 'affected_lines': len(je_blank_user),
        'debit_ec': je_blank_user['debit_amount_ec'].sum(), 'credit_ec': je_blank_user['credit_amount_ec'].sum()
    })

    # 01e: JE Blank Transaction Type
    je_blank_type = gl_clean[gl_clean['transaction_type'] == '']
    dqc_results.append({
        'check': '01e_Warning_JE_Blank_Transaction_Type', 'desc': 'Journal entry data has blank values in Transaction Type',
        'type': 'Warning', 'capability': 'Standard/Non-Standard', 'fields': 'transaction_type',
        'affected_je': je_blank_type['journal_number'].nunique(), 'affected_lines': len(je_blank_type),
        'debit_ec': je_blank_type['debit_amount_ec'].sum(), 'credit_ec': je_blank_type['credit_amount_ec'].sum()
    })

    # 02a: TB Accounts Not in COA
    tb_not_coa = tb_clean[~tb_clean['account_number'].isin(coa_accounts)]
    dqc_results.append({
        'check': '02a_Error_TB_Accounts_Not_In_COA', 'desc': 'Trial balance has accounts not in chart of accounts',
        'type': 'Error', 'capability': 'Core Reconciliation', 'fields': 'account_number',
        'affected_je': 0, 'affected_lines': len(tb_not_coa), 'debit_ec': 0.0, 'credit_ec': 0.0
    })

    # 02b: JE Accounts Not in COA
    je_not_coa = gl_clean[~gl_clean['account_number'].isin(coa_accounts)]
    dqc_results.append({
        'check': '02b_Error_JE_Accounts_Not_In_COA', 'desc': 'Journal entry data has accounts not in chart of accounts',
        'type': 'Error', 'capability': 'Core Reconciliation', 'fields': 'account_number',
        'affected_je': je_not_coa['journal_number'].nunique(), 'affected_lines': len(je_not_coa),
        'debit_ec': je_not_coa['debit_amount_ec'].sum(), 'credit_ec': je_not_coa['credit_amount_ec'].sum()
    })

    # 03a: TB Amount Digits Too Long
    tb_long_digits = tb_clean[tb_clean['ending_balance_ec'].abs() > 1e15]
    dqc_results.append({
        'check': '03a_Error_TB_Amount_Digits_TooLong', 'desc': 'Trial balance amounts have excess digits',
        'type': 'Error', 'capability': 'Data Integrity', 'fields': 'ending_balance_ec',
        'affected_je': 0, 'affected_lines': len(tb_long_digits), 'debit_ec': 0.0, 'credit_ec': 0.0
    })

    # 03b: JE Amount Digits Too Long
    je_long_digits = gl_clean[gl_clean['net_amount_ec'].abs() > 1e15]
    dqc_results.append({
        'check': '03b_Error_JE_Amount_Digits_TooLong', 'desc': 'Journal entry amounts have excess digits',
        'type': 'Error', 'capability': 'Data Integrity', 'fields': 'net_amount_ec',
        'affected_je': je_long_digits['journal_number'].nunique(), 'affected_lines': len(je_long_digits),
        'debit_ec': je_long_digits['debit_amount_ec'].sum(), 'credit_ec': je_long_digits['credit_amount_ec'].sum()
    })

    # 04a: COA Duplicate Account Numbers
    coa_dup = coa_clean[coa_clean.duplicated(subset=['account_number'], keep=False)]
    dqc_results.append({
        'check': '04a_Error_COA_Duplicate_Account_Numbers', 'desc': 'Chart of accounts has duplicate account numbers',
        'type': 'Error', 'capability': 'COA Integrity', 'fields': 'account_number',
        'affected_je': 0, 'affected_lines': len(coa_dup), 'debit_ec': 0.0, 'credit_ec': 0.0
    })

    # 04b: TB Duplicate Account Numbers
    tb_dup = tb_clean[tb_clean.duplicated(subset=['entity_id', 'Fiscal_Year_Identifier', 'account_number'], keep=False)]
    dqc_results.append({
        'check': '04b_Error_TB_Duplicate_Account_Numbers', 'desc': 'Trial balance has duplicate account numbers within the same fiscal period',
        'type': 'Error', 'capability': 'TB Integrity', 'fields': 'account_number',
        'affected_je': 0, 'affected_lines': len(tb_dup), 'debit_ec': 0.0, 'credit_ec': 0.0
    })

    # 05: JE Unknown Classification
    je_unknown_class = gl_clean[~gl_clean['is_standard'].isin(['S', 'N'])]
    dqc_results.append({
        'check': '05_Error_JE_Unknown_Classification', 'desc': 'Journal entry has lines with unknown classifications (standard/nonstandard)',
        'type': 'Error', 'capability': 'Classification', 'fields': 'is_standard',
        'affected_je': je_unknown_class['journal_number'].nunique(), 'affected_lines': len(je_unknown_class),
        'debit_ec': je_unknown_class['debit_amount_ec'].sum(), 'credit_ec': je_unknown_class['credit_amount_ec'].sum()
    })

    # 06: JE Multiple Classification
    je_multi_class_docs = gl_clean.groupby('journal_number')['is_standard'].nunique()
    multi_class_jn_set = set(je_multi_class_docs[je_multi_class_docs > 1].index)
    je_multi_class = gl_clean[gl_clean['journal_number'].isin(multi_class_jn_set)]
    dqc_results.append({
        'check': '06_Error_JE_Multiple_Classification', 'desc': 'Journal entry has lines with multiple classifications in single entry',
        'type': 'Error', 'capability': 'Classification', 'fields': 'journal_number, is_standard',
        'affected_je': len(multi_class_jn_set), 'affected_lines': len(je_multi_class),
        'debit_ec': je_multi_class['debit_amount_ec'].sum(), 'credit_ec': je_multi_class['credit_amount_ec'].sum()
    })

    # 07: COA Unknown Financial Statement Category
    valid_fs_cats = {'Assets', 'Liabilities', 'Equity', 'Revenue', 'Expenses'}
    coa_bad_cat = coa_clean[~coa_clean['financial_statement_category'].isin(valid_fs_cats)]
    dqc_results.append({
        'check': '07_Error_COA_Unknown_Financial_Statement_Category', 'desc': 'Chart of accounts has accounts with an unknown financial statement category',
        'type': 'Error', 'capability': 'FS Presentation', 'fields': 'financial_statement_category',
        'affected_je': 0, 'affected_lines': len(coa_bad_cat), 'debit_ec': 0.0, 'credit_ec': 0.0
    })

    # 08: JE Sum of Amount by Entry Not Net Zero
    doc_lines_cnt = gl_clean.groupby('journal_number').size()
    multi_line_docs = set(doc_lines_cnt[doc_lines_cnt > 1].index)
    doc_sums = gl_clean[gl_clean['journal_number'].isin(multi_line_docs)].groupby('journal_number')['net_amount_ec'].sum()
    unbalanced_je_set = set(doc_sums[doc_sums.abs() > 0.01].index)
    je_unbalanced = gl_clean[gl_clean['journal_number'].isin(unbalanced_je_set)]
    dqc_results.append({
        'check': '08_Warning_JE_Sum_of_Amount_by_Entry_Not_Net_Zero', 'desc': 'Journal entry data has multi-line entries that do not net to zero',
        'type': 'Warning', 'capability': 'Double Entry Control', 'fields': 'journal_number, net_amount_ec',
        'affected_je': len(unbalanced_je_set), 'affected_lines': len(je_unbalanced),
        'debit_ec': je_unbalanced['debit_amount_ec'].sum(), 'credit_ec': je_unbalanced['credit_amount_ec'].sum()
    })

    # 09: JE One Line Entries
    one_line_docs = set(doc_lines_cnt[doc_lines_cnt == 1].index)
    je_one_line = gl_clean[gl_clean['journal_number'].isin(one_line_docs)]
    dqc_results.append({
        'check': '09_Warning_JE_One_Line_Entries', 'desc': 'Journal entry has only one line',
        'type': 'Warning', 'capability': 'Double Entry Control', 'fields': 'journal_number',
        'affected_je': len(one_line_docs), 'affected_lines': len(je_one_line),
        'debit_ec': je_one_line['debit_amount_ec'].sum(), 'credit_ec': je_one_line['credit_amount_ec'].sum()
    })

    # 10: JE Entry Amount Consistency
    je_inconsistent = gl_clean[(gl_clean['net_amount_ec'] - (gl_clean['debit_amount_ec'] - gl_clean['credit_amount_ec'])).abs() > 0.01]
    dqc_results.append({
        'check': '10_Warning_JE_Entry_Amount_Consistency', 'desc': 'Journal entry has lines with net amount inconsistent with debit/credit',
        'type': 'Warning', 'capability': 'Math Verification', 'fields': 'net_amount_ec, debit_amount_ec, credit_amount_ec',
        'affected_je': je_inconsistent['journal_number'].nunique(), 'affected_lines': len(je_inconsistent),
        'debit_ec': je_inconsistent['debit_amount_ec'].sum(), 'credit_ec': je_inconsistent['credit_amount_ec'].sum()
    })

    # 11: JE Debit Credit Same Line
    je_both_dr_cr = gl_clean[(gl_clean['debit_amount_ec'] > 0) & (gl_clean['credit_amount_ec'] > 0)]
    dqc_results.append({
        'check': '11_Warning_JE_Debit_Credit_Same_Line', 'desc': 'Journal entry has lines with both non-zero debit and credit amounts',
        'type': 'Warning', 'capability': 'Math Verification', 'fields': 'debit_amount_ec, credit_amount_ec',
        'affected_je': je_both_dr_cr['journal_number'].nunique(), 'affected_lines': len(je_both_dr_cr),
        'debit_ec': je_both_dr_cr['debit_amount_ec'].sum(), 'credit_ec': je_both_dr_cr['credit_amount_ec'].sum()
    })

    # 12: JE Amount Currency Inconsistency
    je_curr_incons = gl_clean[((gl_clean['net_amount_ec'] == 0) & (gl_clean['net_amount_gc'] != 0)) | ((gl_clean['net_amount_ec'] != 0) & (gl_clean['net_amount_gc'] == 0) & (gl_clean['group_currency_gc'] != ''))]
    dqc_results.append({
        'check': '12_Warning_JE_Amount_Currency_Inconsistency', 'desc': 'Journal entry has lines with zero in one currency and non-zero in another',
        'type': 'Warning', 'capability': 'Multi-Currency', 'fields': 'net_amount_ec, net_amount_gc',
        'affected_je': je_curr_incons['journal_number'].nunique(), 'affected_lines': len(je_curr_incons),
        'debit_ec': je_curr_incons['debit_amount_ec'].sum(), 'credit_ec': je_curr_incons['credit_amount_ec'].sum()
    })

    # 13a: JE Entity Multiple Currency
    ent_curr_cnt = gl_clean.groupby('entity_id')['entity_currency_ec'].nunique()
    multi_curr_ents = set(ent_curr_cnt[ent_curr_cnt > 1].index)
    je_multi_ent_curr = gl_clean[gl_clean['entity_id'].isin(multi_curr_ents)]
    dqc_results.append({
        'check': '13a_Warning_JE_Entity_Multiple_Currency', 'desc': 'Journal entry data has individual entities with multiple entity currencies',
        'type': 'Warning', 'capability': 'Multi-Currency', 'fields': 'entity_id, entity_currency_ec',
        'affected_je': je_multi_ent_curr['journal_number'].nunique(), 'affected_lines': len(je_multi_ent_curr),
        'debit_ec': je_multi_ent_curr['debit_amount_ec'].sum(), 'credit_ec': je_multi_ent_curr['credit_amount_ec'].sum()
    })

    # 13b: JE Group Multiple Currency
    grp_currs = gl_clean[gl_clean['group_currency_gc'] != '']['group_currency_gc'].nunique()
    dqc_results.append({
        'check': '13b_Warning_JE_Group_Multiple_Currency', 'desc': 'Journal entry data has multiple group currencies',
        'type': 'Warning', 'capability': 'Multi-Currency', 'fields': 'group_currency_gc',
        'affected_je': gl_clean['journal_number'].nunique() if grp_currs > 1 else 0,
        'affected_lines': len(gl_clean) if grp_currs > 1 else 0,
        'debit_ec': gl_clean['debit_amount_ec'].sum() if grp_currs > 1 else 0.0,
        'credit_ec': gl_clean['credit_amount_ec'].sum() if grp_currs > 1 else 0.0
    })

    # 14: JE Multiple Date Values
    doc_dates_cnt = gl_clean.groupby('journal_number')['date_effective'].nunique()
    multi_date_docs = set(doc_dates_cnt[doc_dates_cnt > 1].index)
    je_multi_dates = gl_clean[gl_clean['journal_number'].isin(multi_date_docs)]
    dqc_results.append({
        'check': '14_Warning_JE_Multiple_Date_Values', 'desc': 'Journal entry has lines with multiple effective or posted dates',
        'type': 'Warning', 'capability': 'Temporal Consistency', 'fields': 'journal_number, date_effective',
        'affected_je': len(multi_date_docs), 'affected_lines': len(je_multi_dates),
        'debit_ec': je_multi_dates['debit_amount_ec'].sum(), 'credit_ec': je_multi_dates['credit_amount_ec'].sum()
    })

    # 15: JE Multiple Transaction Types
    doc_types_cnt = gl_clean.groupby('journal_number')['transaction_type'].nunique()
    multi_type_docs = set(doc_types_cnt[doc_types_cnt > 1].index)
    je_multi_type = gl_clean[gl_clean['journal_number'].isin(multi_type_docs)]
    dqc_results.append({
        'check': '15_Warning_JE_Multiple_Transaction_Type', 'desc': 'Journal entry has lines with multiple transaction types',
        'type': 'Warning', 'capability': 'Transaction Analytics', 'fields': 'journal_number, transaction_type',
        'affected_je': len(multi_type_docs), 'affected_lines': len(je_multi_type),
        'debit_ec': je_multi_type['debit_amount_ec'].sum(), 'credit_ec': je_multi_type['credit_amount_ec'].sum()
    })

    # 16: JE Prior/Post Effective Date
    tp_start = date_to_iso(params.get('testingPeriodStart', '2025-04-01'))
    tp_end = date_to_iso(params.get('testingPeriodEnd', '2026-03-31'))
    gl_iso_dates = gl_clean['date_effective'].apply(date_to_iso)
    if tp_start and tp_end:
        out_of_period = gl_iso_dates.apply(lambda d: d is not None and (d < tp_start or d > tp_end))
    else:
        out_of_period = pd.Series([False] * len(gl_clean))
    je_out_period = gl_clean[out_of_period]
    dqc_results.append({
        'check': '16_Warning_JE_Prior_Post_Effective_Date', 'desc': 'Journal entry has lines with effective dates prior or after testing period',
        'type': 'Warning', 'capability': 'Period Cutoff', 'fields': 'date_effective',
        'affected_je': je_out_period['journal_number'].nunique(), 'affected_lines': len(je_out_period),
        'debit_ec': je_out_period['debit_amount_ec'].sum(), 'credit_ec': je_out_period['credit_amount_ec'].sum()
    })

    # 17: JE Multiple User ID
    doc_users_cnt = gl_clean.groupby('journal_number')['userid_entered'].nunique()
    multi_user_docs = set(doc_users_cnt[doc_users_cnt > 1].index)
    je_multi_users = gl_clean[gl_clean['journal_number'].isin(multi_user_docs)]
    dqc_results.append({
        'check': '17_Observation_JE_Multiple_User_ID', 'desc': 'Journal entry has lines with multiple User IDs in single document',
        'type': 'Observation', 'capability': 'User Analytics', 'fields': 'journal_number, userid_entered',
        'affected_je': len(multi_user_docs), 'affected_lines': len(je_multi_users),
        'debit_ec': je_multi_users['debit_amount_ec'].sum(), 'credit_ec': je_multi_users['credit_amount_ec'].sum()
    })

    # 18: JE Multiple Entry Description
    doc_desc_cnt = gl_clean.groupby('journal_number')['journal_header_description'].nunique()
    multi_desc_docs = set(doc_desc_cnt[doc_desc_cnt > 1].index)
    je_multi_desc = gl_clean[gl_clean['journal_number'].isin(multi_desc_docs)]
    dqc_results.append({
        'check': '18_Observation_JE_Multiple_Entry_Description', 'desc': 'Journal entry has multiple entry header descriptions',
        'type': 'Observation', 'capability': 'Narration Analysis', 'fields': 'journal_number, journal_header_description',
        'affected_je': len(multi_desc_docs), 'affected_lines': len(je_multi_desc),
        'debit_ec': je_multi_desc['debit_amount_ec'].sum(), 'credit_ec': je_multi_desc['credit_amount_ec'].sum()
    })

    # 19: JE Sum of Amount by Transaction Type Not Net Zero
    tt_sums = gl_clean.groupby('transaction_type')['net_amount_ec'].sum()
    unbalanced_tt = set(tt_sums[tt_sums.abs() > 1.0].index)
    je_unbal_tt = gl_clean[gl_clean['transaction_type'].isin(unbalanced_tt)]
    dqc_results.append({
        'check': '19_Observation_JE_Sum_of_Amount_by_Transaction_Type_Not_Net_Zero', 'desc': 'Journal entry data has transaction types that do not net to zero',
        'type': 'Observation', 'capability': 'Transaction Analytics', 'fields': 'transaction_type, net_amount_ec',
        'affected_je': je_unbal_tt['journal_number'].nunique(), 'affected_lines': len(je_unbal_tt),
        'debit_ec': je_unbal_tt['debit_amount_ec'].sum(), 'credit_ec': je_unbal_tt['credit_amount_ec'].sum()
    })

    # 20: UserID Entered Multiple User Name Entered
    user_names_cnt = gl_clean[gl_clean['userid_entered'] != ''].groupby('userid_entered')['user_name_entered'].nunique()
    multi_name_users = set(user_names_cnt[user_names_cnt > 1].index)
    je_multi_name_users = gl_clean[gl_clean['userid_entered'].isin(multi_name_users)]
    dqc_results.append({
        'check': '20_Observation_UserID_Entered_Multiple_User_Name_Entered', 'desc': 'User ID is mapped to more than one user name value',
        'type': 'Observation', 'capability': 'User Analytics', 'fields': 'userid_entered, user_name_entered',
        'affected_je': je_multi_name_users['journal_number'].nunique(), 'affected_lines': len(je_multi_name_users),
        'debit_ec': je_multi_name_users['debit_amount_ec'].sum(), 'credit_ec': je_multi_name_users['credit_amount_ec'].sum()
    })

    def save_dqc_file(df, check_name):
        filename = f'Parquet_Data_Integrity_Check_{check_name}.csv'
        p = os.path.join(output_dir, filename)
        if df is not None and len(df) > 0:
            df.to_csv(p, index=False)
        elif df is not None:
            df.head(0).to_csv(p, index=False)
        else:
            pd.DataFrame().to_csv(p, index=False)

    save_dqc_file(coa_blank, '01a_Error_COA_Blank_Values')
    save_dqc_file(tb_blank, '01b_Error_TB_Blank_Values')
    save_dqc_file(je_blank, '01c_Error_JE_Blank_Values')
    save_dqc_file(je_blank_user, '01d_Warning_JE_Blank_UserID')
    save_dqc_file(je_blank_type, '01e_Warning_JE_Blank_Transaction_Type')
    save_dqc_file(tb_not_coa, '02a_Error_TB_Accounts_Not_In_COA')
    save_dqc_file(je_not_coa, '02b_Error_JE_Accounts_Not_In_COA')
    save_dqc_file(tb_long_digits, '03a_Error_TB_Amount_Digits_TooLong')
    save_dqc_file(je_long_digits, '03b_Error_JE_Amount_Digits_TooLong')
    save_dqc_file(coa_dup, '04a_Error_COA_Duplicate_Account_Numbers')
    save_dqc_file(tb_dup, '04b_Error_TB_Duplicate_Account_Numbers')
    save_dqc_file(je_unknown_class, '05_Error_JE_Unknown_Classification')
    save_dqc_file(je_multi_class, '06_Error_JE_Multiple_Classification')
    save_dqc_file(coa_bad_cat, '07_Error_COA_Unknown_Financial_Statement_Category')
    save_dqc_file(je_unbalanced, '08_Warning_JE_Sum_of_Amount_by_Entry_Not_Net_Zero')
    save_dqc_file(je_one_line, '09_Warning_JE_One_Line_Entries')
    save_dqc_file(je_inconsistent, '10_Warning_JE_Entry_Amount_Consistency')
    save_dqc_file(je_both_dr_cr, '11_Warning_JE_Debit_Credit_Same_Line')
    save_dqc_file(je_curr_incons, '12_Warning_JE_Amount_Currency_Inconsistency')
    save_dqc_file(je_multi_ent_curr, '13a_Warning_JE_Entity_Multiple_Currency')
    save_dqc_file(gl_clean[gl_clean['group_currency_gc'] != ''] if grp_currs > 1 else gl_clean.head(0), '13b_Warning_JE_Group_Multiple_Currency')
    save_dqc_file(je_multi_dates, '14_Warning_JE_Multiple_Date_Values')
    save_dqc_file(je_multi_type, '15_Warning_JE_Multiple_Transaction_Type')
    save_dqc_file(je_out_period, '16_Warning_JE_Prior_Post_Effective_Date')
    save_dqc_file(je_multi_users, '17_Observation_JE_Multiple_User_ID')
    save_dqc_file(je_multi_desc, '18_Observation_JE_Multiple_Entry_Description')
    save_dqc_file(je_unbal_tt, '19_Observation_JE_Sum_of_Amount_by_Transaction_Type_Not_Net_Zero')
    save_dqc_file(je_multi_name_users, '20_Observation_UserID_Entered_Multiple_User_Name_Entered')

    dqc_summary_df = pd.DataFrame(dqc_results)
    dqc_summary_df.columns = [
        'Data_Integrity_Check_Name', 'Description', 'Error_Warning', 'Capability_Impacted',
        'Related_Fields', 'Number_of_Affected_Journal_Entries', 'Number_of_Affected_Lines',
        'Debit_Amount_EC', 'Credit_Amount_EC'
    ]
    dqc_summary_df['Data_Integrity_Check_Toggled_Off'] = 'No'

    dqc_out_csv = os.path.join(output_dir, 'Parquet_Data_Integrity_Check_00_Summary.csv')
    dqc_summary_df.to_csv(dqc_out_csv, index=False)

    # -------------------------------------------------------------
    # 6. REFINE DATA (POPULATION EXCLUSIONS)
    # -------------------------------------------------------------
    log_event(run_id, 'REFINE_DATA', 68, 'Applying Omnia population refinement and exclusions filters', log_file)

    exclusions_config = params.get('exclusions', {}) or {}
    exclude_zeros = exclusions_config.get('excludeZeroLines', params.get('excludeZeroLines', True))
    excluded_sys_types = set([str(x).strip().upper() for x in exclusions_config.get('systemEntryTypes', []) if str(x).strip()])
    excluded_accounts = set([str(x).strip() for x in exclusions_config.get('excludedAccounts', []) if str(x).strip()])
    excluded_entry_types = set([str(x).strip().upper() for x in exclusions_config.get('excludedEntryTypes', []) if str(x).strip()])
    excluded_users = set([str(x).strip().upper() for x in exclusions_config.get('excludedUsers', []) if str(x).strip()])

    total_raw_lines = len(gl_clean)
    excluded_masks = {}
    
    # 1. Zero Value lines
    if exclude_zeros:
        excluded_masks['Zero_Amount'] = (gl_clean['net_amount_ec'] == 0) & (gl_clean['debit_amount_ec'] == 0) & (gl_clean['credit_amount_ec'] == 0)
    else:
        excluded_masks['Zero_Amount'] = pd.Series([False] * total_raw_lines, index=gl_clean.index)

    # 2. System / Recurring entries
    if excluded_sys_types:
        excluded_masks['System_Recurring_Entries'] = gl_clean['transaction_type'].str.upper().isin(excluded_sys_types)
    else:
        excluded_masks['System_Recurring_Entries'] = pd.Series([False] * total_raw_lines, index=gl_clean.index)

    # 3. Specific Accounts
    if excluded_accounts:
        excluded_masks['Specific_Accounts'] = gl_clean['account_number'].isin(excluded_accounts)
    else:
        excluded_masks['Specific_Accounts'] = pd.Series([False] * total_raw_lines, index=gl_clean.index)

    # 4. Entry Types
    if excluded_entry_types:
        excluded_masks['Entry_Types'] = gl_clean['transaction_type'].str.upper().isin(excluded_entry_types)
    else:
        excluded_masks['Entry_Types'] = pd.Series([False] * total_raw_lines, index=gl_clean.index)

    # 5. User Exclusions
    if excluded_users:
        excluded_masks['User_Exclusions'] = gl_clean['userid_entered'].str.upper().isin(excluded_users)
    else:
        excluded_masks['User_Exclusions'] = pd.Series([False] * total_raw_lines, index=gl_clean.index)

    # Combine all exclusions
    is_excluded = pd.Series(False, index=gl_clean.index)
    exclusion_reason = pd.Series('', index=gl_clean.index, dtype=str)

    for reason, mask in excluded_masks.items():
        mask_series = pd.Series(mask, index=gl_clean.index)
        is_excluded = is_excluded | mask_series
        exclusion_reason = pd.Series(
            np.where((exclusion_reason == '') & mask_series, reason, exclusion_reason),
            index=gl_clean.index
        )

    gl_refined = gl_clean[~is_excluded].copy().reset_index(drop=True)
    gl_excluded = gl_clean[is_excluded].copy().reset_index(drop=True)
    gl_excluded['exclusion_reason'] = exclusion_reason[is_excluded].to_numpy()

    # Save Refined Population and Exclusions Summary
    refined_out_csv = os.path.join(output_dir, 'Refined_Population.csv')
    gl_refined.to_csv(refined_out_csv, index=False)

    exclusions_summary_records = [
        {'Exclusion_Category': 'Zero-Value Lines ($0 Amount)', 'Excluded_Line_Count': int(excluded_masks['Zero_Amount'].sum()), 'Audit_Rationale': exclusions_config.get('rationales', {}).get('zeroLines', 'Omit zero dollar header/balancing placeholder lines')},
        {'Exclusion_Category': 'System or Recurring Entries', 'Excluded_Line_Count': int(excluded_masks['System_Recurring_Entries'].sum()), 'Audit_Rationale': exclusions_config.get('rationales', {}).get('systemEntries', 'Standard automated month-end depreciation/amortization')},
        {'Exclusion_Category': 'Specific Routine Accounts', 'Excluded_Line_Count': int(excluded_masks['Specific_Accounts'].sum()), 'Audit_Rationale': exclusions_config.get('rationales', {}).get('accounts', 'Intercompany clearing and treasury routine accounts')},
        {'Exclusion_Category': 'Specific Entry Types', 'Excluded_Line_Count': int(excluded_masks['Entry_Types'].sum()), 'Audit_Rationale': exclusions_config.get('rationales', {}).get('entryTypes', 'Automated system clearing transaction types')},
        {'Exclusion_Category': 'Automated System Users', 'Excluded_Line_Count': int(excluded_masks['User_Exclusions'].sum()), 'Audit_Rationale': exclusions_config.get('rationales', {}).get('users', 'Pre-approved batch automation accounts')},
    ]
    exclusions_summary_df = pd.DataFrame(exclusions_summary_records)
    exclusions_summary_df['Total_Raw_Population'] = total_raw_lines
    exclusions_summary_df['Refined_Testing_Population'] = len(gl_refined)

    exclusions_out_csv = os.path.join(output_dir, 'Omnia_Exclusions_Summary.csv')
    exclusions_summary_df.to_csv(exclusions_out_csv, index=False)

    # -------------------------------------------------------------
    # 7. OMNIA PARAMETRIC FRAUD & ANOMALY TESTS (ON REFINED DATA)
    # -------------------------------------------------------------
    log_event(run_id, 'PARAMETER_TESTS', 76, 'Running Omnia JE Parametric Fraud Analytics Suite on refined population', log_file)

    tests_config = params.get('testsConfig', {}) or {}
    flagged_entries_map = {} # document_number -> dict of details and set of flagged tests
    param_summary = {}

    def flag_entry(doc_no, test_name, reason_desc, line_row=None, severity='MEDIUM'):
        if not doc_no:
            return
        if doc_no not in flagged_entries_map:
            flagged_entries_map[doc_no] = {
                'journal_number': doc_no,
                'entity_id': line_row['entity_id'] if line_row is not None else '',
                'date_effective': line_row['date_effective'] if line_row is not None else '',
                'date_posted': line_row['date_posted'] if line_row is not None else '',
                'account_number': line_row['account_number'] if line_row is not None else '',
                'account_description': line_row['account_description'] if line_row is not None else '',
                'net_amount_ec': float(line_row['net_amount_ec']) if line_row is not None else 0.0,
                'entity_currency_ec': line_row['entity_currency_ec'] if line_row is not None else 'INR',
                'userid_entered': line_row['userid_entered'] if line_row is not None else '',
                'user_name_entered': line_row['user_name_entered'] if line_row is not None else '',
                'transaction_type': line_row['transaction_type'] if line_row is not None else '',
                'journal_header_description': line_row['journal_header_description'] if line_row is not None else '',
                'journal_line_description': line_row['journal_line_description'] if line_row is not None else '',
                'flagged_tests': [],
                'reasons': [],
                'severity': severity
            }
        if test_name not in flagged_entries_map[doc_no]['flagged_tests']:
            flagged_entries_map[doc_no]['flagged_tests'].append(test_name)
            flagged_entries_map[doc_no]['reasons'].append(reason_desc)

    # TEST 1: Seldom Used Accounts (Expected to Run - Mandatory)
    cfg_seldom = tests_config.get('seldomAccounts', {})
    if cfg_seldom.get('enabled', True):
        thresh_cnt = int(cfg_seldom.get('thresholdCount', 5))
        amt_thresh = float(cfg_seldom.get('threshold', 0.0))
        custom_seldom = set([str(x).strip() for x in cfg_seldom.get('customAccounts', []) if str(x).strip()])
        
        gl_acc_counts = gl_clean.groupby('account_number')['journal_number'].nunique()
        seldom_accounts = set(gl_acc_counts[(gl_acc_counts >= 1) & (gl_acc_counts <= thresh_cnt)].index)
        seldom_accounts.update(custom_seldom)

        matched_seldom = gl_refined[gl_refined['account_number'].isin(seldom_accounts) & (gl_refined['net_amount_ec'].abs() >= amt_thresh)].copy()
        matched_seldom['Flagged_Test'] = 'Seldom Used Accounts'
        matched_seldom['Flag_Reason'] = matched_seldom['account_number'].apply(lambda a: f"Account {a} posted <= {thresh_cnt} times in period")
        
        seldom_out_csv = os.path.join(output_dir, 'Omnia_Test_Seldom_Accounts.csv')
        matched_seldom.to_csv(seldom_out_csv, index=False)
        param_summary['Seldom_Used_Accounts'] = matched_seldom['journal_number'].nunique()

        for _, row in matched_seldom.iterrows():
            flag_entry(row['journal_number'], 'Seldom Used Accounts', f"Account {row['account_number']} is seldom used", row, 'MEDIUM')
    else:
        param_summary['Seldom_Used_Accounts'] = 0

    # TEST 2: Keywords Scan (Expected to Run - Mandatory)
    cfg_kw = tests_config.get('keywords', {})
    if cfg_kw.get('enabled', True):
        kw_list = cfg_kw.get('keywordList', [
            'plug', 'test', 'fictitious', 'reverse', 'manual', 'bribe', 'fraud', 'conceal',
            'adjustment', 'mistake', 'misstatement', 'officer', 'prize', 'abuse', 'alter',
            'seizure', 'bury', 'corrupt', 'demand', 'embezzle', 'theft', 'suspense', 'net to zero', 'round'
        ])
        kw_thresh = float(cfg_kw.get('threshold', 0.0))
        kw_pattern = r'\b(?:' + '|'.join([re.escape(k.lower().strip()) for k in kw_list if k.strip()]) + r')\b'

        h_match = gl_refined['journal_header_description'].str.lower().str.contains(kw_pattern, regex=True, na=False)
        l_match = gl_refined['journal_line_description'].str.lower().str.contains(kw_pattern, regex=True, na=False)
        amt_match = gl_refined['net_amount_ec'].abs() >= kw_thresh

        matched_kw = gl_refined[(h_match | l_match) & amt_match].copy()
        matched_kw['Flagged_Test'] = 'Suspect Keywords'
        matched_kw['Flag_Reason'] = 'Journal contains suspect fraud keyword terms in narration'

        kw_out_csv = os.path.join(output_dir, 'Omnia_Test_Keywords.csv')
        matched_kw.to_csv(kw_out_csv, index=False)
        param_summary['Keywords_Scan'] = matched_kw['journal_number'].nunique()

        for _, row in matched_kw.iterrows():
            flag_entry(row['journal_number'], 'Suspect Keywords', 'Suspect keyword in description', row, 'HIGH')
    else:
        param_summary['Keywords_Scan'] = 0

    # TEST 3: Closing / Post-Closing Entries (Expected to Run - Mandatory)
    cfg_close = tests_config.get('closingEntries', {})
    if cfg_close.get('enabled', True):
        close_ref_date_str = cfg_close.get('closingDate') or params.get('fiscalYearEnd') or testing_period_end
        close_ref_iso = date_to_iso(close_ref_date_str) or '2026-03-31'
        days_after = int(cfg_close.get('daysAfter', 10))
        days_before = int(cfg_close.get('daysBefore', 0))
        close_thresh = float(cfg_close.get('threshold', 0.0))

        try:
            ref_dt = datetime.date.fromisoformat(close_ref_iso)
            win_start = (ref_dt - datetime.timedelta(days=days_before)).strftime('%Y-%m-%d')
            win_end = (ref_dt + datetime.timedelta(days=days_after)).strftime('%Y-%m-%d')

            gl_ref_dates = gl_refined['date_effective'].apply(date_to_iso)
            in_close_win = gl_ref_dates.apply(lambda d: d is not None and win_start <= d <= win_end)
            amt_match = gl_refined['net_amount_ec'].abs() >= close_thresh

            matched_close = gl_refined[in_close_win & amt_match].copy()
            matched_close['Flagged_Test'] = 'Closing/Post-Closing Entries'
            matched_close['Flag_Reason'] = matched_close['date_effective'].apply(lambda d: f"Dated {d} in post-closing window ({win_start} to {win_end})")

            close_out_csv = os.path.join(output_dir, 'Omnia_Test_Closing_Entries.csv')
            matched_close.to_csv(close_out_csv, index=False)
            param_summary['Closing_Entries'] = matched_close['journal_number'].nunique()

            for _, row in matched_close.iterrows():
                flag_entry(row['journal_number'], 'Closing/Post-Closing Entries', f"Post-close adjustment on {row['date_effective']}", row, 'HIGH')
        except Exception as e:
            matched_close = gl_refined.head(0)
            param_summary['Closing_Entries'] = 0
    else:
        param_summary['Closing_Entries'] = 0

    # TEST 4: Unusual Accounts (Expected to Consider - Optional)
    cfg_unusual = tests_config.get('unusualAccounts', {})
    if cfg_unusual.get('enabled', True):
        unusual_thresh = int(cfg_unusual.get('thresholdCount', 3))
        unusual_amt_thresh = float(cfg_unusual.get('threshold', 0.0))
        
        tb_beg_accounts = set(tb_clean[tb_clean['Fiscal_Year_Identifier'] == 'Current Period Beginning']['account_number'].unique())
        gl_counts = gl_clean.groupby('account_number')['journal_number'].nunique()
        new_or_rare = set(gl_counts[(gl_counts <= unusual_thresh) | (~gl_counts.index.isin(tb_beg_accounts))].index)

        matched_unusual = gl_refined[gl_refined['account_number'].isin(new_or_rare) & (gl_refined['net_amount_ec'].abs() >= unusual_amt_thresh)].copy()
        matched_unusual['Flagged_Test'] = 'Unusual Accounts'
        matched_unusual['Flag_Reason'] = matched_unusual['account_number'].apply(lambda a: f"Account {a} is new or used significantly less than peer baseline")

        unusual_out_csv = os.path.join(output_dir, 'Omnia_Test_Unusual_Accounts.csv')
        matched_unusual.to_csv(unusual_out_csv, index=False)
        param_summary['Unusual_Accounts'] = matched_unusual['journal_number'].nunique()

        for _, row in matched_unusual.iterrows():
            flag_entry(row['journal_number'], 'Unusual Accounts', f"Account {row['account_number']} is unusual vs baseline", row, 'MEDIUM')
    else:
        param_summary['Unusual_Accounts'] = 0

    # TEST 5: Round Amounts & Recurring Digits (Expected to Consider - Optional)
    cfg_round = tests_config.get('roundAmounts', {})
    if cfg_round.get('enabled', True):
        round_multiples = [int(x) for x in cfg_round.get('roundMultiples', [1000, 10000, 100000, 1000000]) if str(x).isdigit()]
        recurring_digits = [int(x) for x in cfg_round.get('recurringDigits', [3, 4, 5]) if str(x).isdigit()]
        round_thresh = float(cfg_round.get('threshold', 0.0))

        def check_round_digit(amt):
            abs_a = abs(int(amt))
            if abs_a == 0 or abs_a < round_thresh:
                return False, ''
            for mult in sorted(round_multiples, reverse=True):
                if abs_a % mult == 0 and abs_a >= mult:
                    return True, f"Multiple of {mult:,}"
            s = str(abs_a)
            for d_len in recurring_digits:
                if len(s) >= d_len:
                    sub = s[-d_len:]
                    if len(set(sub)) == 1:
                        return True, f"Recurring ending digits '{sub}'"
            return False, ''

        round_checks = gl_refined['net_amount_ec'].apply(check_round_digit)
        is_round = round_checks.apply(lambda x: x[0])
        round_reasons = round_checks.apply(lambda x: x[1])

        matched_round = gl_refined[is_round].copy()
        matched_round['Flagged_Test'] = 'Round Amounts & Recurring Digits'
        matched_round['Flag_Reason'] = round_reasons[is_round].to_numpy()

        round_out_csv = os.path.join(output_dir, 'Omnia_Test_Round_Amounts.csv')
        matched_round.to_csv(round_out_csv, index=False)
        param_summary['Round_Amounts'] = matched_round['journal_number'].nunique()

        for _, row in matched_round.iterrows():
            flag_entry(row['journal_number'], 'Round Amounts & Recurring Digits', f"Round transaction: {row['Flag_Reason']}", row, 'MEDIUM')
    else:
        param_summary['Round_Amounts'] = 0

    # TEST 6: Duplicate Entries (Additional / Supplementary)
    cfg_dup = tests_config.get('duplicateEntries', {})
    if cfg_dup.get('enabled', True):
        dup_cnt_thresh = int(cfg_dup.get('countThreshold', 2))
        dup_amt_thresh = float(cfg_dup.get('amountThreshold', 0.0))

        gl_refined['dup_key'] = gl_refined['date_effective'] + '_' + gl_refined['account_number'] + '_' + gl_refined['net_amount_ec'].abs().round(2).astype(str)
        dup_groups = gl_refined[gl_refined['net_amount_ec'].abs() >= dup_amt_thresh].groupby('dup_key')['journal_number'].nunique()
        flagged_dup_keys = set(dup_groups[dup_groups >= dup_cnt_thresh].index)

        matched_dup = gl_refined[gl_refined['dup_key'].isin(flagged_dup_keys)].copy().drop(columns=['dup_key'])
        matched_dup['Flagged_Test'] = 'Duplicate Entries'
        matched_dup['Flag_Reason'] = 'Identical date, account, and amount across multiple entries'

        dup_out_csv = os.path.join(output_dir, 'Omnia_Test_Duplicate_Entries.csv')
        matched_dup.to_csv(dup_out_csv, index=False)
        param_summary['Duplicate_Entries'] = matched_dup['journal_number'].nunique()

        for _, row in matched_dup.iterrows():
            flag_entry(row['journal_number'], 'Duplicate Entries', 'Duplicate date/account/amount combination', row, 'HIGH')
    else:
        param_summary['Duplicate_Entries'] = 0

    # TEST 7: Dates of Interest & Weekend Postings (Additional / Supplementary)
    cfg_dates = tests_config.get('datesOfInterest', {})
    if cfg_dates.get('enabled', True):
        special_dates_iso = set([date_to_iso(d) for d in cfg_dates.get('dates', ['2025-12-25', '2025-12-31', '2026-01-01', '2026-03-31']) if date_to_iso(d)])
        check_weekends = cfg_dates.get('checkWeekends', True)
        dates_thresh = float(cfg_dates.get('threshold', 0.0))

        def is_date_of_interest(d_str):
            if not d_str: return False, ''
            iso = date_to_iso(d_str)
            if not iso: return False, ''
            if iso in special_dates_iso:
                return True, f"Special audit holiday/cutoff date ({iso})"
            if check_weekends:
                try:
                    dt = datetime.date.fromisoformat(iso)
                    if dt.weekday() in (5, 6): # Saturday, Sunday
                        day_name = 'Saturday' if dt.weekday() == 5 else 'Sunday'
                        return True, f"Weekend posting on {day_name} ({iso})"
                except:
                    pass
            return False, ''

        date_checks = gl_refined['date_effective'].apply(is_date_of_interest)
        is_doi = date_checks.apply(lambda x: x[0]) & (gl_refined['net_amount_ec'].abs() >= dates_thresh)
        doi_reasons = date_checks.apply(lambda x: x[1])

        matched_doi = gl_refined[is_doi].copy()
        matched_doi['Flagged_Test'] = 'Dates of Interest'
        matched_doi['Flag_Reason'] = doi_reasons[is_doi].to_numpy()

        doi_out_csv = os.path.join(output_dir, 'Omnia_Test_Dates_Of_Interest.csv')
        matched_doi.to_csv(doi_out_csv, index=False)
        param_summary['Dates_Of_Interest'] = matched_doi['journal_number'].nunique()

        for _, row in matched_doi.iterrows():
            flag_entry(row['journal_number'], 'Dates of Interest', f"Unusual date: {row['Flag_Reason']}", row, 'MEDIUM')
    else:
        param_summary['Dates_Of_Interest'] = 0

    # TEST 8: Debits to Revenue Accounts (Additional / Supplementary)
    cfg_rev = tests_config.get('debitsToRevenue', {})
    if cfg_rev.get('enabled', True):
        custom_rev = set([str(x).strip() for x in cfg_rev.get('revenueAccounts', []) if str(x).strip()])
        if not custom_rev:
            revenue_coa_accounts = set(coa_clean[coa_clean['financial_statement_category'] == 'Revenue']['account_number'].unique())
        else:
            revenue_coa_accounts = custom_rev
        rev_thresh = float(cfg_rev.get('threshold', 0.0))

        is_rev_debit = gl_refined['account_number'].isin(revenue_coa_accounts) & ((gl_refined['net_amount_ec'] > rev_thresh) | (gl_refined['debit_amount_ec'] > rev_thresh))
        matched_rev = gl_refined[is_rev_debit].copy()
        matched_rev['Flagged_Test'] = 'Debits to Revenue'
        matched_rev['Flag_Reason'] = matched_rev['account_number'].apply(lambda a: f"Abnormal debit posting to revenue account {a}")

        rev_out_csv = os.path.join(output_dir, 'Omnia_Test_Debits_To_Revenue.csv')
        matched_rev.to_csv(rev_out_csv, index=False)
        param_summary['Debits_To_Revenue'] = matched_rev['journal_number'].nunique()

        for _, row in matched_rev.iterrows():
            flag_entry(row['journal_number'], 'Debits to Revenue', f"Abnormal debit to revenue account {row['account_number']}", row, 'HIGH')
    else:
        param_summary['Debits_To_Revenue'] = 0

    # TEST 9: Users of Interest & Few Postings Users (Additional / Supplementary)
    cfg_users = tests_config.get('usersOfInterest', {})
    if cfg_users.get('enabled', True):
        monitored_users = set([str(x).strip().upper() for x in cfg_users.get('userList', ['ADMIN', 'SYSTEM', 'BATCH', 'SBPATIL', 'SUPERUSER']) if str(x).strip()])
        few_thresh = int(cfg_users.get('fewPostingsThreshold', 2))
        user_thresh = float(cfg_users.get('threshold', 0.0))

        user_counts = gl_clean.groupby('userid_entered')['journal_number'].nunique()
        few_post_users = set(user_counts[(user_counts > 0) & (user_counts <= few_thresh)].index.str.upper())

        is_user_flagged = (
            gl_refined['userid_entered'].str.upper().isin(monitored_users) |
            gl_refined['userid_entered'].str.upper().isin(few_post_users)
        ) & (gl_refined['net_amount_ec'].abs() >= user_thresh)

        matched_users = gl_refined[is_user_flagged].copy()
        matched_users['Flagged_Test'] = 'Users of Interest'
        matched_users['Flag_Reason'] = matched_users['userid_entered'].apply(
            lambda u: f"High-risk monitored user '{u}'" if u.upper() in monitored_users else f"User '{u}' has rare posting history (<= {few_thresh} entries)"
        )

        users_out_csv = os.path.join(output_dir, 'Omnia_Test_Users_Of_Interest.csv')
        matched_users.to_csv(users_out_csv, index=False)
        param_summary['Users_Of_Interest'] = matched_users['journal_number'].nunique()

        for _, row in matched_users.iterrows():
            flag_entry(row['journal_number'], 'Users of Interest', f"Flagged user: {row['userid_entered']}", row, 'MEDIUM')
    else:
        param_summary['Users_Of_Interest'] = 0

    # TEST 10: BENFORD'S LAW ANALYSIS (FIRST-DIGIT CONFORMITY)
    log_event(run_id, 'BENFORD_ANALYSIS', 84, "Executing Benford's Law first-digit distribution analysis", log_file)

    benford_expected = {
        1: 0.301, 2: 0.176, 3: 0.125, 4: 0.097,
        5: 0.079, 6: 0.067, 7: 0.058, 8: 0.051, 9: 0.046
    }
    
    pos_amts = gl_refined[gl_refined['net_amount_ec'].abs() > 0]['net_amount_ec'].abs()
    first_digits = []
    for a in pos_amts:
        s = f"{a:.6f}".replace('.', '').lstrip('0')
        if s and s[0].isdigit() and s[0] != '0':
            first_digits.append(int(s[0]))

    total_benford_trans = len(first_digits)
    benford_digit_stats = []
    mad_sum = 0.0

    if total_benford_trans > 0:
        digit_counts = pd.Series(first_digits).value_counts().to_dict()
        for d in range(1, 10):
            cnt = digit_counts.get(d, 0)
            actual_pct = round(cnt / total_benford_trans, 4)
            exp_pct = benford_expected[d]
            diff_pct = round(actual_pct - exp_pct, 4)
            mad_sum += abs(diff_pct)
            is_anomaly = abs(diff_pct) > 0.05
            benford_digit_stats.append({
                'digit': d,
                'count': cnt,
                'actualPct': round(actual_pct * 100, 2),
                'expectedPct': round(exp_pct * 100, 2),
                'diffPct': round(diff_pct * 100, 2),
                'isAnomaly': is_anomaly
            })
        mad = mad_sum / 9.0
        conformity_score = max(0.0, min(100.0, round((1.0 - (mad / 0.08)) * 100, 1)))
        if conformity_score >= 85:
            conf_level = 'HIGH'
        elif conformity_score >= 65:
            conf_level = 'ACCEPTABLE'
        else:
            conf_level = 'NON_CONFORMING'
    else:
        conformity_score = 100.0
        conf_level = 'HIGH'
        for d in range(1, 10):
            benford_digit_stats.append({
                'digit': d, 'count': 0, 'actualPct': round(benford_expected[d]*100, 2),
                'expectedPct': round(benford_expected[d]*100, 2), 'diffPct': 0.0, 'isAnomaly': False
            })

    benford_df = pd.DataFrame(benford_digit_stats)
    benford_df.columns = ['First_Digit', 'Transaction_Count', 'Actual_Frequency_Pct', 'Expected_Benford_Pct', 'Variance_Pct', 'Anomaly_Flag']
    benford_out_csv = os.path.join(output_dir, 'Omnia_Benford_Analysis.csv')
    benford_df.to_csv(benford_out_csv, index=False)

    benford_summary = {
        'conformityScore': conformity_score,
        'conformityLevel': conf_level,
        'totalTransactionsTested': total_benford_trans,
        'firstDigitDistribution': benford_digit_stats
    }

    # TEST 11: CONTROL SAMPLING (REPRODUCIBLE RANDOM DUMP)
    log_event(run_id, 'CONTROL_SAMPLE', 88, 'Generating reproducible audit control sample', log_file)

    sample_size = int(params.get('controlSampleCount', 40))
    unique_refined_docs = pd.Series(gl_refined['journal_number'].unique())
    n_sample = min(len(unique_refined_docs), sample_size)
    sampled_docs = set(unique_refined_docs.sample(n=n_sample, random_state=42).tolist()) if n_sample > 0 else set()

    sample_df = gl_refined[gl_refined['journal_number'].isin(sampled_docs)].copy()
    sample_df['Control_Sample'] = 'Sample_Selected'
    sample_out_csv = os.path.join(output_dir, 'Omnia_Control_Sample.csv')
    sample_df.to_csv(sample_out_csv, index=False)

    # -------------------------------------------------------------
    # 8. CONSOLIDATING FLAGGED ENTRIES & TICKMARK REFINEMENT
    # -------------------------------------------------------------
    log_event(run_id, 'CONSOLIDATION', 90, 'Consolidating flagged exceptions and evaluating risk levels', log_file)

    tickmarks_list = params.get('tickmarks', []) or []
    tickmarked_entry_ids = set()
    for tm in tickmarks_list:
        for eid in tm.get('entryIds', []):
            tickmarked_entry_ids.add(str(eid).strip())

    flagged_rows = []
    for doc_no, item in flagged_entries_map.items():
        test_count = len(item['flagged_tests'])
        tests_str = ', '.join(item['flagged_tests'])
        reasons_str = '; '.join(item['reasons'])
        
        if test_count >= 3 or ('Debits to Revenue' in item['flagged_tests'] and 'Suspect Keywords' in item['flagged_tests']):
            risk = 'HIGH'
        elif test_count == 2 or 'Seldom Used Accounts' in item['flagged_tests'] or abs(item['net_amount_ec']) >= 100000:
            risk = 'MEDIUM'
        else:
            risk = 'LOW'

        is_tickmarked = doc_no in tickmarked_entry_ids
        
        flagged_rows.append({
            'Journal_Number': doc_no,
            'Entity_ID': item['entity_id'],
            'Date_Effective': item['date_effective'],
            'Date_Posted': item['date_posted'],
            'Account_Number': item['account_number'],
            'Account_Description': item['account_description'],
            'Net_Amount_EC': item['net_amount_ec'],
            'Entity_Currency': item['entity_currency_ec'],
            'User_ID': item['userid_entered'],
            'User_Name': item['user_name_entered'],
            'Transaction_Type': item['transaction_type'],
            'Flagged_Tests_Count': test_count,
            'Flagged_Tests': tests_str,
            'Flag_Reasons': reasons_str,
            'Risk_Level': risk,
            'Tickmarked_Status': 'Resolved via Tickmark' if is_tickmarked else 'Flagged for Review',
            'Header_Description': item['journal_header_description'],
            'Line_Description': item['journal_line_description']
        })

    flagged_df = pd.DataFrame(flagged_rows)
    if len(flagged_df) > 0:
        flagged_df = flagged_df.sort_values(by=['Risk_Level', 'Flagged_Tests_Count', 'Net_Amount_EC'], ascending=[True, False, False])
    flagged_all_csv = os.path.join(output_dir, 'Omnia_Flagged_Entries_All.csv')
    flagged_df.to_csv(flagged_all_csv, index=False)

    tickmark_records = []
    for tm in tickmarks_list:
        tickmark_records.append({
            'Tickmark_ID': tm.get('id', ''),
            'Title': tm.get('title', ''),
            'Audit_Rationale': tm.get('rationale', ''),
            'Associated_Entries_Count': len(tm.get('entryIds', [])),
            'Send_For_Evaluation': 'Yes' if tm.get('sendForEvaluation') else 'No',
            'Created_By': tm.get('createdBy', 'Auditor'),
            'Created_Date': tm.get('createdDate', datetime.datetime.now().strftime('%Y-%m-%d'))
        })
    tickmark_df = pd.DataFrame(tickmark_records)
    tickmark_out_csv = os.path.join(output_dir, 'Omnia_Tickmark_Summary.csv')
    tickmark_df.to_csv(tickmark_out_csv, index=False)

    total_flagged_count = len(flagged_df)
    high_risk_count = int((flagged_df['Risk_Level'] == 'HIGH').sum()) if len(flagged_df) > 0 else 0
    med_risk_count = int((flagged_df['Risk_Level'] == 'MEDIUM').sum()) if len(flagged_df) > 0 else 0
    low_risk_count = int((flagged_df['Risk_Level'] == 'LOW').sum()) if len(flagged_df) > 0 else 0
    clean_entries_count = max(0, gl_refined['journal_number'].nunique() - total_flagged_count)

    # -------------------------------------------------------------
    # 9. CONTROL TOTALS & STRATIFICATION
    # -------------------------------------------------------------
    log_event(run_id, 'CONTROL_TOTALS', 93, 'Generating Control Totals and Line Stratification', log_file)

    ct_period = gl_clean.groupby('fiscal_period').agg(
        line_count=('journal_line_number', 'count'),
        debit_total=('debit_amount_ec', 'sum'),
        credit_total=('credit_amount_ec', 'sum'),
        net_total=('net_amount_ec', 'sum'),
        date_min=('date_effective', 'min'),
        date_max=('date_effective', 'max')
    ).reset_index()
    ct_period.to_csv(os.path.join(output_dir, 'Control_Total_By_Period.csv'), index=False)

    ct_std = gl_clean.groupby('is_standard').agg(
        line_count=('journal_line_number', 'count'),
        debit_total=('debit_amount_ec', 'sum'),
        credit_total=('credit_amount_ec', 'sum'),
        net_total=('net_amount_ec', 'sum'),
        date_min=('date_effective', 'min'),
        date_max=('date_effective', 'max')
    ).reset_index()
    ct_std.to_csv(os.path.join(output_dir, 'Control_Total_By_Standard_Non_Standard.csv'), index=False)

    ct_curr = gl_clean.groupby('entity_currency_ec').agg(
        line_count=('journal_line_number', 'count'),
        debit_total=('debit_amount_ec', 'sum'),
        credit_total=('credit_amount_ec', 'sum'),
        net_total=('net_amount_ec', 'sum'),
        date_min=('date_effective', 'min'),
        date_max=('date_effective', 'max')
    ).reset_index()
    ct_curr.to_csv(os.path.join(output_dir, 'Control_Total_By_Currency.csv'), index=False)

    ct_tt = gl_clean.groupby('transaction_type').agg(
        line_count=('journal_line_number', 'count'),
        debit_total=('debit_amount_ec', 'sum'),
        credit_total=('credit_amount_ec', 'sum'),
        net_total=('net_amount_ec', 'sum'),
        date_min=('date_effective', 'min'),
        date_max=('date_effective', 'max')
    ).reset_index()
    ct_tt.to_csv(os.path.join(output_dir, 'Control_Total_By_Transaction_Type.csv'), index=False)

    ct_user = gl_clean.groupby('userid_entered').agg(
        line_count=('journal_line_number', 'count'),
        debit_total=('debit_amount_ec', 'sum'),
        credit_total=('credit_amount_ec', 'sum'),
        net_total=('net_amount_ec', 'sum'),
        date_min=('date_effective', 'min'),
        date_max=('date_effective', 'max')
    ).reset_index()
    ct_user.to_csv(os.path.join(output_dir, 'Control_Total_By_User.csv'), index=False)

    doc_lines = gl_clean.groupby(['entity_id', 'is_standard', 'journal_number']).size().reset_index(name='line_count')
    def bucket_line_count(cnt):
        if cnt == 1: return '1'
        if 2 <= cnt <= 20: return '2 - 20'
        if 21 <= cnt <= 100: return '21 - 100'
        if 101 <= cnt <= 1000: return '101 - 1000'
        return '> 1000'

    doc_lines['Number_of_Lines_per_JE'] = doc_lines['line_count'].apply(bucket_line_count)
    strat = doc_lines.groupby(['entity_id', 'is_standard', 'Number_of_Lines_per_JE']).agg(
        Number_of_JE=('journal_number', 'count'),
        Number_of_Lines=('line_count', 'sum')
    ).reset_index()

    strat['Percentage_of_Lines'] = (strat['Number_of_Lines'] / len(gl_clean) * 100).round(2)
    strat['Percentage_of_Entries'] = (strat['Number_of_JE'] / gl_clean['journal_number'].nunique() * 100).round(2)
    strat.to_csv(os.path.join(output_dir, 'JE_Line_Distribution.csv'), index=False)

    # -------------------------------------------------------------
    # 10. EXCEL MASTER AUDIT WORKPAPER GENERATION
    # -------------------------------------------------------------
    log_event(run_id, 'EXCEL_EXPORT', 95, 'Building comprehensive multi-tab Omnia JE Audit Workpaper Workbook', log_file)

    excel_template_path = os.path.join(output_dir, 'JE-Recon-and-DIC-Template.xlsx')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='007680', end_color='007680', fill_type='solid') # Deloitte Teal
    border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    def format_sheet(ws, title, df, max_rows=1000):
        ws.title = title
        headers = list(df.columns)
        ws.append(headers)
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for r_idx, row in enumerate(df.head(max_rows).itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border_thin
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    # 1. Summary Sheet
    ws_summary = wb.create_sheet(title='Summary')
    summary_data_list = [
        ['Omnia Audit Engagement Summary Metric', 'Value'],
        ['Engagement Run ID', run_id],
        ['Workflow Engine', 'Deloitte Omnia JE (Reconciliation & Parametric Fraud Suite)'],
        ['Fiscal Year Tested', params.get('fiscalYear', 2026)],
        ['Testing Period Start', testing_period_start],
        ['Testing Period End', testing_period_end],
        ['Total Input Raw JE Lines', len(gl_clean)],
        ['Excluded Non-Fraud Lines', len(gl_excluded)],
        ['Refined Testing Population Lines', len(gl_refined)],
        ['Total Journal Entries Tested', gl_refined['journal_number'].nunique()],
        ['Total Flagged Exceptions', total_flagged_count],
        ['High Risk Exceptions', high_risk_count],
        ['Medium Risk Exceptions', med_risk_count],
        ['Low Risk Exceptions', low_risk_count],
        ["Benford's Law Conformity Score", f"{conformity_score}% ({conf_level})"],
        ['Total TB Accounts', len(tb_clean)],
        ['Total Master COA Accounts', len(coa_clean)],
        ['Reconciled Accounts (TB vs JE)', reconciled_count],
        ['Unreconciled Accounts (TB vs JE)', unreconciled_count],
        ['DQC Critical Errors Flagged', int((dqc_summary_df['Error_Warning'] == 'Error').sum())],
        ['DQC Warnings Flagged', int((dqc_summary_df['Error_Warning'] == 'Warning').sum())],
        ['DQC Observations Flagged', int((dqc_summary_df['Error_Warning'] == 'Observation').sum())],
    ]
    for r_idx, row in enumerate(summary_data_list, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.border = border_thin
                if c_idx == 1:
                    cell.font = Font(name='Calibri', size=11, bold=True)

    # 2. Flagged Entries Tab
    ws_flags = wb.create_sheet(title='Flagged Exceptions')
    format_sheet(ws_flags, 'Flagged Exceptions', flagged_df)

    # 3. Benford's Law Analysis Tab
    ws_benford = wb.create_sheet(title="Benford's Law")
    format_sheet(ws_benford, "Benford's Law", benford_df)

    # 4. Reconciliation Tab
    ws_recon = wb.create_sheet(title='Reconciliation')
    format_sheet(ws_recon, 'Reconciliation', recon_df[['entity_id', 'account_number', 'account_description', 'financial_statement_category', 'ending_balance', 'beginning_balance', 'je_activity', 'trial_activity', 'variance', 'abs_variance', 'reconciliation']])

    # 5. DQC Summary Tab
    ws_dqc = wb.create_sheet(title='20 DQC Matrix')
    format_sheet(ws_dqc, '20 DQC Matrix', dqc_summary_df)

    # 6. Exclusions Summary Tab
    ws_excl = wb.create_sheet(title='Population Exclusions')
    format_sheet(ws_excl, 'Population Exclusions', exclusions_summary_df)

    # 7. Control Totals Tab
    ws_ct = wb.create_sheet(title='Control Totals')
    format_sheet(ws_ct, 'Control Totals', ct_period)

    # 8. JE Line Distribution Tab
    ws_strat = wb.create_sheet(title='Line Stratification')
    format_sheet(ws_strat, 'Line Stratification', strat)

    wb.save(excel_template_path)

    # -------------------------------------------------------------
    # 11. OUTPUT MANIFEST & SUMMARY JSON
    # -------------------------------------------------------------
    log_event(run_id, 'OUTPUT_GENERATION', 98, 'Compiling Omnia output manifest and summary indicators', log_file)

    outputs = [
        {'id': 'excel_template', 'name': 'JE-Recon-and-DIC-Template.xlsx', 'type': 'xlsx', 'category': 'MASTER', 'description': 'Enterprise Omnia Master Audit Workpaper & Reconciliation Workbook', 'rowCount': len(flagged_df), 'path': excel_template_path},
        {'id': 'flagged_all', 'name': 'Omnia_Flagged_Entries_All.csv', 'type': 'csv', 'category': 'ANOMALY', 'description': 'Consolidated Flagged Journal Entries with Risk Categorization', 'rowCount': len(flagged_df), 'path': flagged_all_csv},
        {'id': 'benford_csv', 'name': 'Omnia_Benford_Analysis.csv', 'type': 'csv', 'category': 'ANOMALY', 'description': "Benford's Law First-Digit Conformity Distribution", 'rowCount': len(benford_df), 'path': benford_out_csv},
        {'id': 'excl_summary', 'name': 'Omnia_Exclusions_Summary.csv', 'type': 'csv', 'category': 'EXCLUSION', 'description': 'Population Refinement and Exclusions Summary', 'rowCount': len(exclusions_summary_df), 'path': exclusions_out_csv},
        {'id': 'refined_pop', 'name': 'Refined_Population.csv', 'type': 'csv', 'category': 'MASTER', 'description': 'Refined Journal Entry Testing Population', 'rowCount': len(gl_refined), 'path': refined_out_csv},
        {'id': 'tickmark_summary', 'name': 'Omnia_Tickmark_Summary.csv', 'type': 'csv', 'category': 'AUDIT_TRAIL', 'description': 'Auditor False-Positive Tickmarks & Resolutions', 'rowCount': len(tickmark_df), 'path': tickmark_out_csv},
        {'id': 'test_seldom', 'name': 'Omnia_Test_Seldom_Accounts.csv', 'type': 'csv', 'category': 'PARAMETER_TEST', 'description': 'Test 1: Seldom Used Accounts Flagged Entries', 'rowCount': param_summary.get('Seldom_Used_Accounts', 0), 'path': os.path.join(output_dir, 'Omnia_Test_Seldom_Accounts.csv')},
        {'id': 'test_kw', 'name': 'Omnia_Test_Keywords.csv', 'type': 'csv', 'category': 'PARAMETER_TEST', 'description': 'Test 2: Suspect Keywords Scan Flagged Entries', 'rowCount': param_summary.get('Keywords_Scan', 0), 'path': os.path.join(output_dir, 'Omnia_Test_Keywords.csv')},
        {'id': 'test_close', 'name': 'Omnia_Test_Closing_Entries.csv', 'type': 'csv', 'category': 'PARAMETER_TEST', 'description': 'Test 3: Post-Closing Period Entries', 'rowCount': param_summary.get('Closing_Entries', 0), 'path': os.path.join(output_dir, 'Omnia_Test_Closing_Entries.csv')},
        {'id': 'test_unusual', 'name': 'Omnia_Test_Unusual_Accounts.csv', 'type': 'csv', 'category': 'PARAMETER_TEST', 'description': 'Test 4: Unusual Accounts Flagged Entries', 'rowCount': param_summary.get('Unusual_Accounts', 0), 'path': os.path.join(output_dir, 'Omnia_Test_Unusual_Accounts.csv')},
        {'id': 'test_round', 'name': 'Omnia_Test_Round_Amounts.csv', 'type': 'csv', 'category': 'PARAMETER_TEST', 'description': 'Test 5: Round Amounts & Recurring Digits', 'rowCount': param_summary.get('Round_Amounts', 0), 'path': os.path.join(output_dir, 'Omnia_Test_Round_Amounts.csv')},
        {'id': 'test_dup', 'name': 'Omnia_Test_Duplicate_Entries.csv', 'type': 'csv', 'category': 'PARAMETER_TEST', 'description': 'Test 6: Duplicate Journal Entries', 'rowCount': param_summary.get('Duplicate_Entries', 0), 'path': os.path.join(output_dir, 'Omnia_Test_Duplicate_Entries.csv')},
        {'id': 'test_doi', 'name': 'Omnia_Test_Dates_Of_Interest.csv', 'type': 'csv', 'category': 'PARAMETER_TEST', 'description': 'Test 7: Dates of Interest & Weekend Postings', 'rowCount': param_summary.get('Dates_Of_Interest', 0), 'path': os.path.join(output_dir, 'Omnia_Test_Dates_Of_Interest.csv')},
        {'id': 'test_rev', 'name': 'Omnia_Test_Debits_To_Revenue.csv', 'type': 'csv', 'category': 'PARAMETER_TEST', 'description': 'Test 8: Debits to Revenue Accounts', 'rowCount': param_summary.get('Debits_To_Revenue', 0), 'path': os.path.join(output_dir, 'Omnia_Test_Debits_To_Revenue.csv')},
        {'id': 'test_users', 'name': 'Omnia_Test_Users_Of_Interest.csv', 'type': 'csv', 'category': 'PARAMETER_TEST', 'description': 'Test 9: Monitored & Rare Users', 'rowCount': param_summary.get('Users_Of_Interest', 0), 'path': os.path.join(output_dir, 'Omnia_Test_Users_Of_Interest.csv')},
        {'id': 'control_sample', 'name': 'Omnia_Control_Sample.csv', 'type': 'csv', 'category': 'CONTROL_SAMPLE', 'description': 'Reproducible Control Sample Dump', 'rowCount': len(sample_df), 'path': sample_out_csv},
        {'id': 'recon_csv', 'name': 'Parquet_Reconciliation.csv', 'type': 'csv', 'category': 'RECONCILIATION', 'description': 'Account Reconciliation Summary (TB vs JE)', 'rowCount': len(recon_df), 'path': recon_out_csv},
        {'id': 'unrecon_csv', 'name': 'Unreconciled_Accounts_Detail.csv', 'type': 'csv', 'category': 'RECONCILIATION', 'description': 'Unreconciled Accounts Detail', 'rowCount': len(unrecon_df), 'path': unrecon_out_csv},
        {'id': 'dqc_summary', 'name': 'Parquet_Data_Integrity_Check_00_Summary.csv', 'type': 'csv', 'category': 'DQC', 'description': 'Omnia 20 Golden Checks (DQC 01a - 20) Summary', 'rowCount': len(dqc_summary_df), 'path': dqc_out_csv},
        {'id': 'ct_period', 'name': 'Control_Total_By_Period.csv', 'type': 'csv', 'category': 'CONTROL_TOTAL', 'description': 'Control Totals grouped by Fiscal Period', 'rowCount': len(ct_period), 'path': os.path.join(output_dir, 'Control_Total_By_Period.csv')},
        {'id': 'ct_std', 'name': 'Control_Total_By_Standard_Non_Standard.csv', 'type': 'csv', 'category': 'CONTROL_TOTAL', 'description': 'Control Totals by Standard/Non-Standard', 'rowCount': len(ct_std), 'path': os.path.join(output_dir, 'Control_Total_By_Standard_Non_Standard.csv')},
        {'id': 'ct_curr', 'name': 'Control_Total_By_Currency.csv', 'type': 'csv', 'category': 'CONTROL_TOTAL', 'description': 'Control Totals by Entity Currency', 'rowCount': len(ct_curr), 'path': os.path.join(output_dir, 'Control_Total_By_Currency.csv')},
        {'id': 'ct_tt', 'name': 'Control_Total_By_Transaction_Type.csv', 'type': 'csv', 'category': 'CONTROL_TOTAL', 'description': 'Control Totals by Transaction Type', 'rowCount': len(ct_tt), 'path': os.path.join(output_dir, 'Control_Total_By_Transaction_Type.csv')},
        {'id': 'ct_user', 'name': 'Control_Total_By_User.csv', 'type': 'csv', 'category': 'CONTROL_TOTAL', 'description': 'Control Totals by User ID', 'rowCount': len(ct_user), 'path': os.path.join(output_dir, 'Control_Total_By_User.csv')},
        {'id': 'strat_csv', 'name': 'JE_Line_Distribution.csv', 'type': 'csv', 'category': 'CONTROL_TOTAL', 'description': 'JE Line Stratification and Distribution', 'rowCount': len(strat), 'path': os.path.join(output_dir, 'JE_Line_Distribution.csv')},
        {'id': 'gl_detail', 'name': 'General_Ledger_Detail.csv', 'type': 'csv', 'category': 'MASTER', 'description': 'Standardized General Ledger Detail (CDM format)', 'rowCount': len(gl_clean), 'path': gl_out_csv},
        {'id': 'tb_detail', 'name': 'Trial_Balance.csv', 'type': 'csv', 'category': 'MASTER', 'description': 'Standardized Trial Balance (CDM format)', 'rowCount': len(tb_clean), 'path': tb_out_csv},
        {'id': 'coa_detail', 'name': 'Chart_of_Accounts.csv', 'type': 'csv', 'category': 'MASTER', 'description': 'Standardized Chart of Accounts (CDM format)', 'rowCount': len(coa_clean), 'path': coa_out_csv},
    ]

    summary_data = {
        'runId': run_id,
        'workflow': 'OMNIA_JET',
        'status': 'COMPLETED',
        'progress': 100,
        'completedAt': datetime.datetime.now().isoformat(),
        'totalInputRows': {
            'tb': len(tb_clean),
            'gl': len(gl_clean),
            'coa': len(coa_clean)
        },
        'exclusionsSummary': {
            'totalInputLines': total_raw_lines,
            'excludedZeroCount': int(excluded_masks['Zero_Amount'].sum()),
            'excludedSystemCount': int(excluded_masks['System_Recurring_Entries'].sum()),
            'excludedAccountsCount': int(excluded_masks['Specific_Accounts'].sum()),
            'excludedUsersCount': int(excluded_masks['User_Exclusions'].sum()),
            'totalExcludedLines': len(gl_excluded),
            'remainingRefinedLines': len(gl_refined)
        },
        'parameterSummary': param_summary,
        'flaggedSummary': {
            'totalFlagged': total_flagged_count,
            'highRiskCount': high_risk_count,
            'medRiskCount': med_risk_count,
            'lowRiskCount': low_risk_count
        },
        'testOutputsSummary': {
            'seldomAccounts': param_summary.get('Seldom_Used_Accounts', 0),
            'keywords': param_summary.get('Keywords_Scan', 0),
            'closingEntries': param_summary.get('Closing_Entries', 0),
            'unusualAccounts': param_summary.get('Unusual_Accounts', 0),
            'roundAmounts': param_summary.get('Round_Amounts', 0),
            'duplicateEntries': param_summary.get('Duplicate_Entries', 0),
            'datesOfInterest': param_summary.get('Dates_Of_Interest', 0),
            'debitsToRevenue': param_summary.get('Debits_To_Revenue', 0),
            'usersOfInterest': param_summary.get('Users_Of_Interest', 0),
        },
        'benfordSummary': benford_summary,
        'controlSampleCount': len(sample_df),
        'riskBreakdown': {
            'highRisk': high_risk_count,
            'mediumRisk': med_risk_count,
            'lowRisk': low_risk_count,
            'cleanEntries': clean_entries_count
        },
        'tickmarkSummary': {
            'totalTickmarks': len(tickmarks_list),
            'totalEntriesResolved': len(tickmarked_entry_ids),
            'totalEntriesPending': max(0, total_flagged_count - len(tickmarked_entry_ids))
        },
        'reconciliationSummary': {
            'totalAccounts': len(recon_df),
            'reconciledAccounts': reconciled_count,
            'unreconciledAccounts': unreconciled_count,
            'totalBeginningBalance': round(float(recon_df['beginning_balance'].sum()), 2),
            'totalEndingBalance': round(float(recon_df['ending_balance'].sum()), 2),
            'totalJEActivity': round(float(recon_df['je_activity'].sum()), 2),
            'totalTrialActivity': round(float(recon_df['trial_activity'].sum()), 2),
            'totalVariance': round(float(recon_df['variance'].sum()), 2)
        },
        'dqcSummary': {
            'totalErrors': int(((dqc_summary_df['Error_Warning'] == 'Error') & ((dqc_summary_df['Number_of_Affected_Lines'] > 0) | (dqc_summary_df['Number_of_Affected_Journal_Entries'] > 0))).sum()),
            'totalWarnings': int(((dqc_summary_df['Error_Warning'] == 'Warning') & ((dqc_summary_df['Number_of_Affected_Lines'] > 0) | (dqc_summary_df['Number_of_Affected_Journal_Entries'] > 0))).sum()),
            'totalObservations': int(((dqc_summary_df['Error_Warning'] == 'Observation') & ((dqc_summary_df['Number_of_Affected_Lines'] > 0) | (dqc_summary_df['Number_of_Affected_Journal_Entries'] > 0))).sum()),
            'checksPassed': int(((dqc_summary_df['Number_of_Affected_Lines'] == 0) & (dqc_summary_df['Number_of_Affected_Journal_Entries'] == 0)).sum()),
            'checksFailed': int(((dqc_summary_df['Number_of_Affected_Lines'] > 0) | (dqc_summary_df['Number_of_Affected_Journal_Entries'] > 0)).sum())
        },
        'controlTotalsSummary': {
            'totalDebit': round(float(gl_clean['debit_amount_ec'].sum()), 2),
            'totalCredit': round(float(gl_clean['credit_amount_ec'].sum()), 2),
            'netAmount': round(float(gl_clean['net_amount_ec'].sum()), 2),
            'periodCount': len(ct_period),
            'userCount': len(ct_user)
        },
        'outputs': outputs
    }

    status_path = os.path.join(os.path.dirname(os.path.dirname(config_path)), 'status.json')
    existing_status = {}
    if os.path.exists(status_path):
        try:
            with open(status_path, 'r', encoding='utf-8') as f:
                existing_status = json.load(f)
        except:
            pass

    started_at = existing_status.get('startedAt') or summary_data.get('completedAt') or datetime.datetime.now().isoformat()
    summary_data['startedAt'] = started_at
    existing_status.update(summary_data)
    summary_data = existing_status

    with open(status_path, 'w', encoding='utf-8') as f:
        json.dump(summary_data, f, indent=2)

    log_event(run_id, 'COMPLETED', 100, f'Omnia JET workflow executed successfully. {total_flagged_count} exceptions identified, {len(outputs)} audit artifacts generated.', log_file)
    print(f"__RESULT__{json.dumps(summary_data)}", flush=True)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--config', required=True, help='Path to run_config.json')
    args = parser.parse_args()
    try:
        run_omnia_jet(args.config)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"__ERROR__{json.dumps({'error': str(e), 'traceback': tb})}", file=sys.stderr, flush=True)
        sys.exit(1)
