import os
import random
import datetime
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_spark_50k_audit_workbook(output_path: str):
    print(f"Generating 50K Spark JET Audit Dataset -> {output_path}...")
    random.seed(42)
    np.random.seed(42)

    # -------------------------------------------------------------
    # 1. GENERATE 650 TRIAL BALANCE ACCOUNTS
    # -------------------------------------------------------------
    tb_accounts = []
    
    # Subtypes distribution:
    # 150 Assets (100000 - 199999)
    # 120 Liabilities (200000 - 299999)
    #  60 Equity (300000 - 399999)
    # 120 Revenue (400000 - 499999)
    # 200 Expenses (500000 - 699999)
    # Total = 650
    
    asset_fs_lines = ['Cash and Cash Equivalents', 'Trade Receivables', 'Prepaid Expenses', 'Inventories', 'Property Plant and Equipment', 'Accumulated Depreciation', 'Intangibles']
    liab_fs_lines = ['Trade Accounts Payable', 'Accrued Compensation', 'Taxes Payable', 'Short-Term Debt', 'Current Liabilities', 'Long-Term Notes Payable']
    equity_fs_lines = ['Common Stock', 'Additional Paid-in Capital', 'Retained Earnings', 'Treasury Stock', 'Accumulated Other Comprehensive Income']
    rev_fs_lines = ['Product Sales Revenue', 'Service Consulting Revenue', 'Subscription Software Revenue', 'License and Royalties', 'Other Operating Revenue']
    exp_fs_lines = ['Cost of Goods Sold', 'Salaries and Wages Expense', 'Rent and Occupancy Expense', 'Depreciation Expense', 'Marketing and Advertising', 'Professional Fees', 'Information Technology Expense', 'Travel and Entertainment Expense', 'Office Supplies Expense']

    # Assets (150)
    for i in range(1, 151):
        gl = f"{100000 + i * 50:06d}"
        fs = asset_fs_lines[i % len(asset_fs_lines)]
        desc = f"{fs} Account #{i:03d}"
        tb_accounts.append({
            'G/L': gl,
            'Description': desc,
            'Account Subtype': 'Assets',
            'FS Line Item': fs,
            'BaseOpening': round(random.uniform(50000, 500000), 2)
        })

    # Liabilities (120)
    for i in range(1, 121):
        gl = f"{200000 + i * 50:06d}"
        fs = liab_fs_lines[i % len(liab_fs_lines)]
        desc = f"{fs} Account #{i:03d}"
        tb_accounts.append({
            'G/L': gl,
            'Description': desc,
            'Account Subtype': 'Liabilities',
            'FS Line Item': fs,
            'BaseOpening': -round(random.uniform(40000, 450000), 2)
        })

    # Equity (60)
    for i in range(1, 61):
        gl = f"{300000 + i * 50:06d}"
        fs = equity_fs_lines[i % len(equity_fs_lines)]
        desc = f"{fs} Account #{i:03d}"
        tb_accounts.append({
            'G/L': gl,
            'Description': desc,
            'Account Subtype': 'Equity',
            'FS Line Item': fs,
            'BaseOpening': -round(random.uniform(20000, 250000), 2)
        })

    # Revenue (120)
    for i in range(1, 121):
        gl = f"{400000 + i * 50:06d}"
        fs = rev_fs_lines[i % len(rev_fs_lines)]
        desc = f"{fs} Account #{i:03d}"
        tb_accounts.append({
            'G/L': gl,
            'Description': desc,
            'Account Subtype': 'Revenue',
            'FS Line Item': fs,
            'BaseOpening': 0.00  # P&L accounts start with 0 opening
        })

    # Expenses (200)
    for i in range(1, 201):
        gl = f"{500000 + i * 50:06d}"
        fs = exp_fs_lines[i % len(exp_fs_lines)]
        desc = f"{fs} Account #{i:03d}"
        tb_accounts.append({
            'G/L': gl,
            'Description': desc,
            'Account Subtype': 'Expenses',
            'FS Line Item': fs,
            'BaseOpening': 0.00  # P&L accounts start with 0 opening
        })

    assert len(tb_accounts) == 650, f"Expected 650 accounts, got {len(tb_accounts)}"

    # Balance Opening Balances exactly to 0.00
    total_opening = sum(a['BaseOpening'] for a in tb_accounts)
    # Adjust on Retained Earnings account
    retained_acc = next(a for a in tb_accounts if 'Retained Earnings' in a['Description'])
    retained_acc['BaseOpening'] = round(retained_acc['BaseOpening'] - total_opening, 2)

    # Double check net zero opening balance
    net_opening = round(sum(a['BaseOpening'] for a in tb_accounts), 2)
    assert abs(net_opening) == 0.0, f"Opening balance not zero: {net_opening}"

    gl_account_list = [a['G/L'] for a in tb_accounts]
    gl_asset_accounts = [a['G/L'] for a in tb_accounts if a['Account Subtype'] == 'Assets']
    gl_liab_accounts = [a['G/L'] for a in tb_accounts if a['Account Subtype'] == 'Liabilities']
    gl_rev_accounts = [a['G/L'] for a in tb_accounts if a['Account Subtype'] == 'Revenue']
    gl_exp_accounts = [a['G/L'] for a in tb_accounts if a['Account Subtype'] == 'Expenses']

    # Tracking account activity for TB Debits & Credits
    account_debits = {gl: 0.0 for gl in gl_account_list}
    account_credits = {gl: 0.0 for gl in gl_account_list}

    # -------------------------------------------------------------
    # 2. GENERATE 50,000 POPULATION (GL) TRANSACTIONS (NET ZERO)
    # -------------------------------------------------------------
    print("Generating 50,000 journal transaction rows...")

    users = ['JSMITH', 'ASIMMONS', 'MWILLIAMS', 'FIN_ADMIN', 'SYSTEM', 'KPATEL', 'EJONES', 'BWHITE', 'RCHEN', 'DGREEN', 'MLEE']
    doc_types = ['SA', 'AB', 'RV', 'KR', 'KZ', 'DZ', 'AA', 'WE']
    headers_text = [
        'Monthly operational expense accrual', 'Customer cash receipt settlement',
        'Vendor invoice standard payment', 'Monthly depreciation run',
        'Direct payroll batch clearing', 'Software license subscription revenue',
        'Inventory valuation adjustment', 'Prepaid insurance amortization',
        'Intercompany trade settlement', 'Travel and expense reimbursement'
    ]

    start_date = datetime.date(2026, 1, 1)
    end_date = datetime.date(2026, 12, 31)
    date_range_days = (end_date - start_date).days

    gl_rows = []
    doc_counter = 10000001

    # Plan:
    # 15,000 2-line entries = 30,000 rows
    #  4,000 3-line entries = 12,000 rows
    #  2,000 4-line entries =  8,000 rows
    # Total = 50,000 rows!

    # 15,000 2-line documents
    for _ in range(15000):
        doc_no = str(doc_counter)
        doc_counter += 1
        d_days = random.randint(0, date_range_days)
        pstng_date = (start_date + datetime.timedelta(days=d_days)).strftime('%Y-%m-%d')
        entry_date = pstng_date
        doc_date = pstng_date
        dtype = random.choice(doc_types)
        user = random.choice(users)
        htext = random.choice(headers_text)

        # Amount
        amt = round(random.uniform(50.0, 45000.0), 2)
        dr_gl = random.choice(gl_exp_accounts + gl_asset_accounts)
        cr_gl = random.choice(gl_rev_accounts + gl_liab_accounts + gl_asset_accounts)
        while cr_gl == dr_gl:
            cr_gl = random.choice(gl_rev_accounts + gl_liab_accounts)

        # Line 1: Debit (+amt)
        gl_rows.append({
            'G/L': dr_gl,
            'DocumentNo': doc_no,
            'Type': dtype,
            'Entry Date': entry_date,
            'Pstng Date': pstng_date,
            'Doc. Date': doc_date,
            'Amount in local cur.': amt,
            'LCurr': 'USD',
            'Amount in doc. curr.': amt,
            'Curr.': 'USD',
            'Amount in loc.curr.2': amt,
            'LCur2': 'USD',
            'Document Header Text': htext,
            'Text': f'Debit clearing for {htext}',
            'User name': user
        })
        account_debits[dr_gl] += amt

        # Line 2: Credit (-amt)
        gl_rows.append({
            'G/L': cr_gl,
            'DocumentNo': doc_no,
            'Type': dtype,
            'Entry Date': entry_date,
            'Pstng Date': pstng_date,
            'Doc. Date': doc_date,
            'Amount in local cur.': -amt,
            'LCurr': 'USD',
            'Amount in doc. curr.': -amt,
            'Curr.': 'USD',
            'Amount in loc.curr.2': -amt,
            'LCur2': 'USD',
            'Document Header Text': htext,
            'Text': f'Credit offset for {htext}',
            'User name': user
        })
        account_credits[cr_gl] += amt

    # 4,000 3-line documents (1 Debit, 2 Credits or 2 Debits, 1 Credit) -> 12,000 rows
    for _ in range(4000):
        doc_no = str(doc_counter)
        doc_counter += 1
        d_days = random.randint(0, date_range_days)
        pstng_date = (start_date + datetime.timedelta(days=d_days)).strftime('%Y-%m-%d')
        dtype = random.choice(['SA', 'AB', 'KR', 'KZ'])
        user = random.choice(users)
        htext = random.choice(headers_text)

        total_amt = round(random.uniform(500.0, 75000.0), 2)
        part1 = round(total_amt * random.uniform(0.3, 0.7), 2)
        part2 = round(total_amt - part1, 2)

        dr_gl = random.choice(gl_exp_accounts + gl_asset_accounts)
        cr_gl1 = random.choice(gl_liab_accounts + gl_asset_accounts)
        cr_gl2 = random.choice(gl_rev_accounts + gl_liab_accounts)

        # Line 1: Debit total
        gl_rows.append({
            'G/L': dr_gl,
            'DocumentNo': doc_no,
            'Type': dtype,
            'Entry Date': pstng_date,
            'Pstng Date': pstng_date,
            'Doc. Date': pstng_date,
            'Amount in local cur.': total_amt,
            'LCurr': 'USD',
            'Amount in doc. curr.': total_amt,
            'Curr.': 'USD',
            'Amount in loc.curr.2': total_amt,
            'LCur2': 'USD',
            'Document Header Text': htext,
            'Text': 'Split journal main line',
            'User name': user
        })
        account_debits[dr_gl] += total_amt

        # Line 2: Credit part 1
        gl_rows.append({
            'G/L': cr_gl1,
            'DocumentNo': doc_no,
            'Type': dtype,
            'Entry Date': pstng_date,
            'Pstng Date': pstng_date,
            'Doc. Date': pstng_date,
            'Amount in local cur.': -part1,
            'LCurr': 'USD',
            'Amount in doc. curr.': -part1,
            'Curr.': 'USD',
            'Amount in loc.curr.2': -part1,
            'LCur2': 'USD',
            'Document Header Text': htext,
            'Text': 'Split offset item 1',
            'User name': user
        })
        account_credits[cr_gl1] += part1

        # Line 3: Credit part 2
        gl_rows.append({
            'G/L': cr_gl2,
            'DocumentNo': doc_no,
            'Type': dtype,
            'Entry Date': pstng_date,
            'Pstng Date': pstng_date,
            'Doc. Date': pstng_date,
            'Amount in local cur.': -part2,
            'LCurr': 'USD',
            'Amount in doc. curr.': -part2,
            'Curr.': 'USD',
            'Amount in loc.curr.2': -part2,
            'LCur2': 'USD',
            'Document Header Text': htext,
            'Text': 'Split offset item 2',
            'User name': user
        })
        account_credits[cr_gl2] += part2

    # 2,000 4-line documents -> 8,000 rows
    for _ in range(2000):
        doc_no = str(doc_counter)
        doc_counter += 1
        d_days = random.randint(0, date_range_days)
        pstng_date = (start_date + datetime.timedelta(days=d_days)).strftime('%Y-%m-%d')
        dtype = random.choice(['SA', 'RV', 'AB'])
        user = random.choice(users)
        htext = random.choice(headers_text)

        dr_total = round(random.uniform(1000.0, 90000.0), 2)
        dr1 = round(dr_total * 0.6, 2)
        dr2 = round(dr_total - dr1, 2)

        cr1 = round(dr_total * 0.45, 2)
        cr2 = round(dr_total - cr1, 2)

        gl_d1 = random.choice(gl_asset_accounts)
        gl_d2 = random.choice(gl_exp_accounts)
        gl_c1 = random.choice(gl_rev_accounts)
        gl_c2 = random.choice(gl_liab_accounts)

        # Debit 1
        gl_rows.append({
            'G/L': gl_d1,
            'DocumentNo': doc_no,
            'Type': dtype,
            'Entry Date': pstng_date,
            'Pstng Date': pstng_date,
            'Doc. Date': pstng_date,
            'Amount in local cur.': dr1,
            'LCurr': 'USD',
            'Amount in doc. curr.': dr1,
            'Curr.': 'USD',
            'Amount in loc.curr.2': dr1,
            'LCur2': 'USD',
            'Document Header Text': htext,
            'Text': 'Multi-line balanced debit 1',
            'User name': user
        })
        account_debits[gl_d1] += dr1

        # Debit 2
        gl_rows.append({
            'G/L': gl_d2,
            'DocumentNo': doc_no,
            'Type': dtype,
            'Entry Date': pstng_date,
            'Pstng Date': pstng_date,
            'Doc. Date': pstng_date,
            'Amount in local cur.': dr2,
            'LCurr': 'USD',
            'Amount in doc. curr.': dr2,
            'Curr.': 'USD',
            'Amount in loc.curr.2': dr2,
            'LCur2': 'USD',
            'Document Header Text': htext,
            'Text': 'Multi-line balanced debit 2',
            'User name': user
        })
        account_debits[gl_d2] += dr2

        # Credit 1
        gl_rows.append({
            'G/L': gl_c1,
            'DocumentNo': doc_no,
            'Type': dtype,
            'Entry Date': pstng_date,
            'Pstng Date': pstng_date,
            'Doc. Date': pstng_date,
            'Amount in local cur.': -cr1,
            'LCurr': 'USD',
            'Amount in doc. curr.': -cr1,
            'Curr.': 'USD',
            'Amount in loc.curr.2': -cr1,
            'LCur2': 'USD',
            'Document Header Text': htext,
            'Text': 'Multi-line balanced credit 1',
            'User name': user
        })
        account_credits[gl_c1] += cr1

        # Credit 2
        gl_rows.append({
            'G/L': gl_c2,
            'DocumentNo': doc_no,
            'Type': dtype,
            'Entry Date': pstng_date,
            'Pstng Date': pstng_date,
            'Doc. Date': pstng_date,
            'Amount in local cur.': -cr2,
            'LCurr': 'USD',
            'Amount in doc. curr.': -cr2,
            'Curr.': 'USD',
            'Amount in loc.curr.2': -cr2,
            'LCur2': 'USD',
            'Document Header Text': htext,
            'Text': 'Multi-line balanced credit 2',
            'User name': user
        })
        account_credits[gl_c2] += cr2

    print(f"Total GL population rows generated: {len(gl_rows)}")
    assert len(gl_rows) == 50000, f"Expected exactly 50,000 GL rows, got {len(gl_rows)}"

    gl_df = pd.DataFrame(gl_rows)
    total_gl_net = round(gl_df['Amount in local cur.'].sum(), 2)
    print(f"Population Total Net Balance: {total_gl_net:,.2f} USD")
    assert abs(total_gl_net) == 0.0, f"Population does not sum to zero: {total_gl_net}"

    # -------------------------------------------------------------
    # 3. FINALIZE 650 TB ROWS (Closing = Opening + Debits - Credits)
    # -------------------------------------------------------------
    tb_final_rows = []
    for a in tb_accounts:
        gl = a['G/L']
        dr = round(account_debits[gl], 2)
        cr = round(account_credits[gl], 2)
        op = round(a['BaseOpening'], 2)
        # For Assets/Expenses, normal balance is Debit (+). For Liab/Equity/Rev, normal balance is Credit (-).
        closing = round(op + dr - cr, 2)
        
        tb_final_rows.append({
            'G/L': gl,
            'Description': a['Description'],
            'Account Subtype': a['Account Subtype'],
            'Opening Balance': op,
            'Debit': dr,
            'Credit': cr,
            'Closing Balance': closing,
            'FS Line Item': a['FS Line Item']
        })

    tb_df = pd.DataFrame(tb_final_rows)
    
    total_tb_opening = round(tb_df['Opening Balance'].sum(), 2)
    total_tb_debit = round(tb_df['Debit'].sum(), 2)
    total_tb_credit = round(tb_df['Credit'].sum(), 2)
    total_tb_closing = round(tb_df['Closing Balance'].sum(), 2)

    print(f"TB Summary (650 accounts):")
    print(f" - Total Opening: {total_tb_opening:,.2f}")
    print(f" - Total Debits:  {total_tb_debit:,.2f}")
    print(f" - Total Credits: {total_tb_credit:,.2f}")
    print(f" - Total Closing: {total_tb_closing:,.2f}")

    assert abs(total_tb_opening) == 0.0, f"TB Opening balance not 0: {total_tb_opening}"
    assert abs(total_tb_closing) == 0.0, f"TB Closing balance not 0: {total_tb_closing}"
    assert abs(total_tb_debit - total_tb_credit) == 0.0, f"TB Debit != Credit: {total_tb_debit} vs {total_tb_credit}"

    # -------------------------------------------------------------
    # 4. PREPARE AUDIT PARAMETERS SHEET
    # -------------------------------------------------------------
    params_data = [
        {
            'Parameter Code': 'ENGAGEMENT_NAME',
            'Parameter Name': 'Audit Engagement Name',
            'Configured Value': 'Deloitte Global JET 2026 Audit',
            'Category': 'Engagement Meta',
            'Audit Purpose': 'Primary engagement identifier for JET audit workpapers'
        },
        {
            'Parameter Code': 'FISCAL_YEAR',
            'Parameter Name': 'Audit Fiscal Year',
            'Configured Value': '2026',
            'Category': 'Period Boundaries',
            'Audit Purpose': 'Applicable fiscal evaluation year'
        },
        {
            'Parameter Code': 'START_DATE',
            'Parameter Name': 'Period Start Date',
            'Configured Value': '2026-01-01',
            'Category': 'Period Boundaries',
            'Audit Purpose': 'Beginning date cutoff for journal entry testing population'
        },
        {
            'Parameter Code': 'END_DATE',
            'Parameter Name': 'Period End Date',
            'Configured Value': '2026-12-31',
            'Category': 'Period Boundaries',
            'Audit Purpose': 'Ending date cutoff for journal entry testing population'
        },
        {
            'Parameter Code': 'FINANCIAL_YEAR_END',
            'Parameter Name': 'Financial Year End Date',
            'Configured Value': '2026-12-31',
            'Category': 'Period Boundaries',
            'Audit Purpose': 'Balance sheet cutoff date for closing entries analysis'
        },
        {
            'Parameter Code': 'CURRENCY_CODE',
            'Parameter Name': 'Functional Entity Currency',
            'Configured Value': 'USD',
            'Category': 'Monetary Rules',
            'Audit Purpose': 'Base reporting currency code'
        },
        {
            'Parameter Code': 'MATERIALITY',
            'Parameter Name': 'Overall Materiality (OM)',
            'Configured Value': '500000',
            'Category': 'Monetary Rules',
            'Audit Purpose': 'Overall audit planning materiality threshold'
        },
        {
            'Parameter Code': 'PERF_MATERIALITY',
            'Parameter Name': 'Performance Materiality (PM)',
            'Configured Value': '375000',
            'Category': 'Monetary Rules',
            'Audit Purpose': 'Performance materiality for sampling and exceptions'
        },
        {
            'Parameter Code': 'TRIVIAL_THRESHOLD',
            'Parameter Name': 'Clearly Trivial Threshold (CTT)',
            'Configured Value': '25000',
            'Category': 'Monetary Rules',
            'Audit Purpose': 'Clearly trivial threshold below which items are immaterial'
        },
        {
            'Parameter Code': 'SELECTED_EXCEPTIONS',
            'Parameter Name': 'Exception Tests Evaluated',
            'Configured Value': '1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12',
            'Category': 'Testing Scope',
            'Audit Purpose': 'Active parameter exceptions to run in the Spark execution engine'
        },
        {
            'Parameter Code': 'USERS_OF_INTEREST',
            'Parameter Name': 'Users of Interest (Ex 05)',
            'Configured Value': 'FIN_ADMIN, SYSTEM, KPATEL',
            'Category': 'Exception Parameters',
            'Audit Purpose': 'High-risk or executive usernames evaluated in Parameter Exception 05'
        },
        {
            'Parameter Code': 'KEYWORDS_LIST',
            'Parameter Name': 'Suspicious Keywords (Ex 10)',
            'Configured Value': 'suspense, reserve, plug, adjust, manual, audit, writeoff, error',
            'Category': 'Exception Parameters',
            'Audit Purpose': 'Keywords scanned in journal descriptions in Parameter Exception 10'
        },
        {
            'Parameter Code': 'ROUND_DIGITS',
            'Parameter Name': 'Round Amount Patterns (Ex 08)',
            'Configured Value': '00, 000, 0000, 00000',
            'Category': 'Exception Parameters',
            'Audit Purpose': 'Trailing numeric zero patterns flagged in Parameter Exception 08'
        },
        {
            'Parameter Code': 'CLOSING_DAYS_BEFORE',
            'Parameter Name': 'Days Before Closing Date (Ex 06)',
            'Configured Value': '5',
            'Category': 'Exception Parameters',
            'Audit Purpose': 'Lead window around year-end for closing entries analysis'
        },
        {
            'Parameter Code': 'CLOSING_DAYS_AFTER',
            'Parameter Name': 'Days After Closing Date (Ex 06)',
            'Configured Value': '15',
            'Category': 'Exception Parameters',
            'Audit Purpose': 'Lag window around year-end for closing entries analysis'
        },
        {
            'Parameter Code': 'DUPLICATE_COUNT_LIMIT',
            'Parameter Name': 'Duplicate Count Threshold (Ex 09)',
            'Configured Value': '3',
            'Category': 'Exception Parameters',
            'Audit Purpose': 'Minimum occurrence frequency for duplicate entries'
        },
        {
            'Parameter Code': 'DUPLICATE_AMT_LIMIT',
            'Parameter Name': 'Duplicate Amount Threshold (Ex 09)',
            'Configured Value': '10000',
            'Category': 'Exception Parameters',
            'Audit Purpose': 'Minimum monetary value for duplicate postings detection'
        },
        {
            'Parameter Code': 'CONTROL_SAMPLE_SIZE',
            'Parameter Name': 'Control Sample Population Count',
            'Configured Value': '25',
            'Category': 'Sampling Rules',
            'Audit Purpose': 'Sample size extracted for substantive audit sample testing'
        }
    ]
    params_df = pd.DataFrame(params_data)

    # -------------------------------------------------------------
    # 5. WRITE MULTI-TAB EXCEL WORKBOOK (Polished Formatting)
    # -------------------------------------------------------------
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        gl_df.to_excel(writer, sheet_name='Population', index=False)
        tb_df.to_excel(writer, sheet_name='Trial_Balance', index=False)
        params_df.to_excel(writer, sheet_name='Audit_Parameters', index=False)

    print(f"Formatting workbook with Deloitte styling -> {output_path}...")
    wb = openpyxl.load_workbook(output_path)

    # Deloitte Colors
    teal_fill = PatternFill(start_color='007680', end_color='007680', fill_type='solid')
    header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    regular_font = Font(name='Segoe UI', size=9, color='0F172A')
    mono_font = Font(name='Consolas', size=9, color='007680')
    number_font = Font(name='Consolas', size=9, color='1E293B')
    border_thin = Side(style='thin', color='CBD5E1')
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        
        # Style Header Row
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = teal_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border
        ws.row_dimensions[1].height = 28

        # Style Data Rows
        for row_idx in range(2, min(ws.max_row + 1, 500)):  # Fast sample format for openpyxl speed
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.font = number_font
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.font = regular_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col[:20]:
                val_str = str(cell.value or '')
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_path)
    file_size_mb = os.path.getsize(output_path) / (1024 * 1024)
    print(f"Workbook successfully saved: {output_path} ({file_size_mb:.2f} MB)")
    print("50,000 GL rows, 650 TB accounts, and Audit Parameters ready!")

if __name__ == '__main__':
    target_wb = os.path.join('sample_data', 'Spark_JET_50K_Population_650_TB_Audit_Workbook.xlsx')
    generate_spark_50k_audit_workbook(target_wb)
